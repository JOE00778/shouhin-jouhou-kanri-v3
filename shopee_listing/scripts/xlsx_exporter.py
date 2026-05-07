#!/usr/bin/env python3
"""Shopee mass upload XLSX 导出器（T-306）

把 T-301 ~ T-305 的输出（SPU + ListingDraft + 类目+属性 + 主图）组装成 Shopee 后台
能直接吃的 mass upload XLSX 文件，按类目分别生成。

5 类目 ID → 模板：
    100629  食品/饮料
    100630  化妆/个护（默认 fallback）
    100632  母婴/营养
    100638  文具/办公
    100640  综合/电器

模板 sheet（"模板"）schema：
    row 0  系统字段名（例 "ps_product_name|1|0"）
    row 1  样例 / 类目 ID / shop ID 行
    row 2  中文名（例 "商品名称"）
    row 3  必填标注（"必填"/"选填"/"依条件必填"）
    row 4  字段说明
    row 5  格式规则

数据行从 row 6 开始。Shopee 后台导入器以 row 0 的「裸字段名」（去掉 |1|0 后缀）作为
列定位，因此我们用 calamine 读出 row 0~5 头部、用 openpyxl 重写一份新 XLSX，
保留同样列顺序。原模板 xlsx 含非法 activePane="" → openpyxl 直接 load_workbook
会抛 ValueError，所以不要直接修改模板。

Module API:
    from xlsx_exporter import export_to_xlsx, ExportContext
    paths = export_to_xlsx(spus, listings, attributes, image_refs, out_dir)

CLI:
    python3 scripts/xlsx_exporter.py \\
        --spu-input examples/spu_input_sample.csv \\
        --drafts-dir drafts/ \\
        --out-dir exports/ \\
        [--mock-drafts]      # drafts/<spu_key>.json 缺失时用占位 ListingDraft
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence

import python_calamine
from openpyxl import Workbook

# ---- sibling imports -------------------------------------------------------- #
_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from parse_spu_input import SPU, Variant, parse_spu_input_csv  # noqa: E402
from listing_generator import ListingDraft  # noqa: E402
from category_mapper import (  # noqa: E402
    VALID_CATEGORIES,
    fill_attributes,
    load_attr_defaults,
    load_category_rules,
    pick_category,
)
from image_finder import ImageRef  # noqa: E402

PROJECT_ROOT = _SCRIPTS_DIR.parent
DEFAULT_TEMPLATE_DIR = PROJECT_ROOT / "shopee上架模板"
DEFAULT_DRAFTS_DIR = PROJECT_ROOT / "drafts"

# Template filenames keyed by category_id (5 templates)
TEMPLATE_FILES: Dict[str, str] = {
    "100629": "Shopee_mass_upload_2026-05-06_100629.xlsx",
    "100630": "Shopee_mass_upload_2026-05-06_100630.xlsx",
    "100632": "Shopee_mass_upload_2026-05-06_100632.xlsx",
    "100638": "Shopee_mass_upload_2026-05-06_100638.xlsx",
    "100640": "Shopee_mass_upload_2026-05-06_100640.xlsx",
}

TEMPLATE_SHEET_NAME = "模板"
HEADER_ROWS = 6  # template prefixes 6 metadata rows before data area


# --------------------------------------------------------------------------- #
# Errors
# --------------------------------------------------------------------------- #


class XlsxExporterError(Exception):
    """Raised on configuration / IO failures."""


# --------------------------------------------------------------------------- #
# Template header loader
# --------------------------------------------------------------------------- #


@dataclass
class TemplateHeader:
    """6 metadata rows + bare field names (row 0 with |x|y suffix stripped)."""

    raw_header: List[str]      # row 0 values (with |1|0 etc.)
    bare_fields: List[str]     # row 0 stripped to plain field names
    metadata_rows: List[List[str]]  # rows 0..5 inclusive (verbatim, list-of-lists)


def _strip_field_suffix(raw: str) -> str:
    """`ps_product_name|1|0` → `ps_product_name`. Empty / None → ''."""
    if raw is None:
        return ""
    s = str(raw)
    return s.split("|", 1)[0]


def load_template_header(category_id: str, *, template_dir: Path = DEFAULT_TEMPLATE_DIR) -> TemplateHeader:
    """Read the first 6 rows of the template's "模板" sheet via calamine.

    Returns the bare field name list and the verbatim metadata rows so we can
    write them straight back into a fresh openpyxl workbook.
    """
    if category_id not in TEMPLATE_FILES:
        raise XlsxExporterError(
            f"category_id {category_id!r} 不在 5 类目 {sorted(TEMPLATE_FILES)} 内"
        )
    path = Path(template_dir) / TEMPLATE_FILES[category_id]
    if not path.exists():
        raise XlsxExporterError(f"模板文件不存在: {path}")

    wb = python_calamine.CalamineWorkbook.from_path(str(path))
    if TEMPLATE_SHEET_NAME not in wb.sheet_names:
        raise XlsxExporterError(
            f"模板文件 {path.name} 缺少 '{TEMPLATE_SHEET_NAME}' sheet"
        )
    sheet = wb.get_sheet_by_name(TEMPLATE_SHEET_NAME)
    rows = sheet.to_python()
    if not rows:
        raise XlsxExporterError(f"模板 '{TEMPLATE_SHEET_NAME}' sheet 为空: {path.name}")

    raw_header = [str(c) if c is not None else "" for c in rows[0]]
    bare = [_strip_field_suffix(c) for c in raw_header]
    metadata = []
    for i in range(min(HEADER_ROWS, len(rows))):
        metadata.append([
            "" if v is None else str(v) for v in rows[i]
        ])
    # Pad metadata up to HEADER_ROWS with blank rows of same width
    while len(metadata) < HEADER_ROWS:
        metadata.append([""] * len(raw_header))

    return TemplateHeader(
        raw_header=raw_header,
        bare_fields=bare,
        metadata_rows=metadata,
    )


# --------------------------------------------------------------------------- #
# Data row builder
# --------------------------------------------------------------------------- #


def _variant_attr_pairs(variant: Variant) -> List[tuple[str, str]]:
    """Return ordered (name, value) pairs for variations.

    Convention: capitalize attr name (color → Color, size → Size). Stable order
    follows insertion order of the dict (Python 3.7+ preserves it). Empty values
    are skipped.
    """
    out: List[tuple[str, str]] = []
    for k, v in variant.variant_attrs.items():
        if v is None:
            continue
        sv = str(v).strip()
        if not sv:
            continue
        name = k.strip()
        if not name:
            continue
        cap = name[0].upper() + name[1:] if name[0].isalpha() else name
        out.append((cap, sv))
    return out


def _make_row(
    *,
    fields: Sequence[str],
    spu: SPU,
    variant: Variant,
    listing: ListingDraft,
    category_id: str,
    attributes: Mapping[str, object],
    image_ref: Optional[ImageRef],
    is_first_variant: bool,
) -> List[object]:
    """Build a data row (list aligned to `fields`).

    Shopee convention for multi-variant SPU: title/description/cover only on the
    first variant row; subsequent rows leave them blank (per template inline doc:
    "对于多型号商品，只需填写第一个型号的商品名称，其他型号留空即可").
    """
    row: List[object] = ["" for _ in fields]

    pairs = _variant_attr_pairs(variant)
    weight_kg: Optional[float] = None
    if variant.weight_grams is not None:
        try:
            weight_kg = round(float(variant.weight_grams) / 1000.0, 4)
        except (TypeError, ValueError):
            weight_kg = None

    for idx, field_name in enumerate(fields):
        # Map field_name → value
        if field_name == "ps_category":
            row[idx] = category_id
        elif field_name == "ps_product_name":
            row[idx] = listing.title if is_first_variant else ""
        elif field_name == "ps_product_description":
            row[idx] = listing.description if is_first_variant else ""
        elif field_name == "ps_sku_parent_short":
            row[idx] = spu.spu_key
        elif field_name == "ps_sku_short":
            row[idx] = variant.jan
        elif field_name == "et_title_variation_integration_no":
            # 整合规格编号 — 同 SPU 的 variant 共享一个值；用 spu_key
            row[idx] = spu.spu_key if pairs else ""
        elif field_name == "et_title_variation_1":
            row[idx] = pairs[0][0] if len(pairs) >= 1 else ""
        elif field_name == "et_title_option_for_variation_1":
            row[idx] = pairs[0][1] if len(pairs) >= 1 else ""
        elif field_name == "et_title_variation_2":
            row[idx] = pairs[1][0] if len(pairs) >= 2 else ""
        elif field_name == "et_title_option_for_variation_2":
            row[idx] = pairs[1][1] if len(pairs) >= 2 else ""
        elif field_name == "ps_weight":
            row[idx] = weight_kg if weight_kg is not None else ""
        elif field_name == "ps_item_cover_image":
            row[idx] = image_ref.url if image_ref else ""
        elif field_name == "ps_brand":
            row[idx] = listing.brand_normalized or ""
        #留空字段（Boss 后台手填 / 后台勾选）
        elif field_name in ("ps_price", "ps_stock", "ps_hs_code", "ps_tax_code"):
            row[idx] = ""
        elif field_name.startswith("channel_id."):
            row[idx] = ""
        elif field_name.startswith("ps_product_global_attribute."):
            attr_id = field_name.split(".", 1)[1]
            val = attributes.get(attr_id)
            row[idx] = "" if val is None else val
        else:
            row[idx] = ""

    return row


# --------------------------------------------------------------------------- #
# Workbook writer
# --------------------------------------------------------------------------- #


def _write_workbook(
    *,
    header: TemplateHeader,
    data_rows: Sequence[Sequence[object]],
    out_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = TEMPLATE_SHEET_NAME

    # Replay 6 metadata rows verbatim (preserve column count)
    for r_idx, meta_row in enumerate(header.metadata_rows, start=1):
        for c_idx, val in enumerate(meta_row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Write data rows starting at row HEADER_ROWS+1
    for d_idx, drow in enumerate(data_rows):
        excel_row = HEADER_ROWS + 1 + d_idx
        for c_idx, val in enumerate(drow, start=1):
            ws.cell(row=excel_row, column=c_idx, value=val)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Public entry: export_to_xlsx
# --------------------------------------------------------------------------- #


@dataclass
class ExportResult:
    category_id: str
    path: Path
    spu_count: int
    row_count: int
    missing_image_jans: List[str]


def export_to_xlsx(
    spus: Sequence[SPU],
    listings: Mapping[str, ListingDraft],
    attributes: Mapping[str, Mapping[str, object]],
    image_refs: Mapping[str, Optional[ImageRef]],
    out_dir: Path,
    *,
    category_assignment: Optional[Mapping[str, str]] = None,
    template_dir: Path = DEFAULT_TEMPLATE_DIR,
    timestamp: Optional[str] = None,
    warn_logger=None,
) -> List[ExportResult]:
    """Group SPUs by category and emit one XLSX per category.

    Args:
        spus: list of SPU objects (T-301 output).
        listings: spu_key → ListingDraft (T-303 output, possibly mock).
        attributes: spu_key → {attr_id: value} (T-304 output).
        image_refs: jan → ImageRef|None (T-305 output).
        out_dir: directory to write `shopee_mass_upload_{cat}_{ts}.xlsx` files.
        category_assignment: optional spu_key → category_id (5 选 1).
            If None, calls pick_category() for each SPU.
        template_dir: where the 5 Shopee templates live.
        timestamp: optional override for file name (default: time.strftime).
        warn_logger: callable(str) for missing-image warnings (defaults to
            stderr print). Can be hooked to 飞书.

    Returns:
        List of ExportResult, one per generated XLSX.
    """
    if warn_logger is None:
        warn_logger = lambda msg: print(f"warning: {msg}", file=sys.stderr)  # noqa: E731

    if timestamp is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")

    # Group SPUs by category
    by_cat: Dict[str, List[SPU]] = {}
    for spu in spus:
        if category_assignment and spu.spu_key in category_assignment:
            cat_id = category_assignment[spu.spu_key]
        else:
            listing = listings.get(spu.spu_key)
            cat_id = pick_category(spu, listing)
        if cat_id not in VALID_CATEGORIES:
            warn_logger(
                f"SPU {spu.spu_key} 类目 {cat_id!r} 不在 5 类目内，回退到 100630"
            )
            cat_id = "100630"
        by_cat.setdefault(cat_id, []).append(spu)

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    results: List[ExportResult] = []
    for cat_id, cat_spus in sorted(by_cat.items()):
        header = load_template_header(cat_id, template_dir=template_dir)
        rows: List[List[object]] = []
        missing_imgs: List[str] = []

        for spu in cat_spus:
            listing = listings.get(spu.spu_key)
            if listing is None:
                warn_logger(
                    f"SPU {spu.spu_key} 没有 ListingDraft，跳过；请先跑 T-303"
                )
                continue
            attrs = attributes.get(spu.spu_key, {}) or {}

            for v_idx, variant in enumerate(spu.variants):
                img_ref = image_refs.get(variant.jan)
                if img_ref is None:
                    missing_imgs.append(variant.jan)
                    warn_logger(
                        f"SPU {spu.spu_key} variant {variant.jan} 缺主图（导出时留空）"
                    )
                row = _make_row(
                    fields=header.bare_fields,
                    spu=spu,
                    variant=variant,
                    listing=listing,
                    category_id=cat_id,
                    attributes=attrs,
                    image_ref=img_ref,
                    is_first_variant=(v_idx == 0),
                )
                rows.append(row)

        out_path = out_dir / f"shopee_mass_upload_{cat_id}_{timestamp}.xlsx"
        _write_workbook(header=header, data_rows=rows, out_path=out_path)
        results.append(
            ExportResult(
                category_id=cat_id,
                path=out_path,
                spu_count=len(cat_spus),
                row_count=len(rows),
                missing_image_jans=missing_imgs,
            )
        )

    return results


# --------------------------------------------------------------------------- #
# CLI helpers — load drafts / images
# --------------------------------------------------------------------------- #


def _load_draft_from_disk(drafts_dir: Path, spu_key: str) -> Optional[ListingDraft]:
    p = drafts_dir / f"{spu_key}.json"
    if not p.exists():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        print(f"warning: 解析 {p} 失败: {e}", file=sys.stderr)
        return None
    return ListingDraft(
        title=data.get("title", ""),
        description=data.get("description", ""),
        key_features=list(data.get("key_features", []) or []),
        how_to_use=list(data.get("how_to_use", []) or []),
        ingredients=data.get("ingredients"),
        spec_json=dict(data.get("spec_json", {}) or {}),
        brand_normalized=data.get("brand_normalized", ""),
        hook=data.get("hook", ""),
        model=data.get("model", ""),
        spu_key=data.get("spu_key", spu_key),
    )


def _make_mock_draft(spu: SPU) -> ListingDraft:
    return ListingDraft(
        title=f"<MOCK> {spu.spu_key} — Direct from Japan",
        description=(
            f"<MOCK description for {spu.spu_key}>\n"
            "[NOTE] This is a placeholder generated by xlsx_exporter --mock-drafts. "
            "Replace by running listing_generator (T-303) with ANTHROPIC_API_KEY set. "
            "Smikie Japan will deliver."
        ),
        key_features=["<mock feature 1>", "<mock feature 2>"],
        how_to_use=["<mock step 1>"],
        ingredients=None,
        spec_json={},
        brand_normalized="",
        hook="Direct from Japan",
        model="mock",
        spu_key=spu.spu_key,
    )


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #


def _build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="xlsx_exporter",
        description="Shopee mass upload XLSX 导出（按类目分文件，T-306）",
    )
    p.add_argument("--spu-input", required=True, help="SPU CSV 路径")
    p.add_argument(
        "--drafts-dir",
        default=str(DEFAULT_DRAFTS_DIR),
        help=f"ListingDraft JSON 目录（默认 {DEFAULT_DRAFTS_DIR}）",
    )
    p.add_argument("--out-dir", required=True, help="XLSX 输出目录")
    p.add_argument(
        "--mock-drafts",
        action="store_true",
        help="drafts/<spu_key>.json 缺失时用 <MOCK> 占位（dry-run 用）",
    )
    p.add_argument(
        "--no-image-fetch",
        action="store_true",
        help="跳过 T-305 主图查找（image_refs 全部为 None；CI/dry-run 用）",
    )
    p.add_argument(
        "--template-dir",
        default=str(DEFAULT_TEMPLATE_DIR),
        help=f"Shopee 模板目录（默认 {DEFAULT_TEMPLATE_DIR}）",
    )
    return p


def main(argv: Optional[List[str]] = None) -> int:
    args = _build_argparser().parse_args(argv)

    spu_path = Path(args.spu_input)
    drafts_dir = Path(args.drafts_dir)
    out_dir = Path(args.out_dir)
    template_dir = Path(args.template_dir)

    try:
        spus = parse_spu_input_csv(spu_path)
    except Exception as e:  # noqa: BLE001
        print(f"error: 解析 SPU CSV 失败: {e}", file=sys.stderr)
        return 2
    if not spus:
        print("error: SPU 输入为空", file=sys.stderr)
        return 2

    # Drafts
    listings: Dict[str, ListingDraft] = {}
    missing_drafts: List[str] = []
    for spu in spus:
        draft = _load_draft_from_disk(drafts_dir, spu.spu_key)
        if draft is None:
            if args.mock_drafts:
                draft = _make_mock_draft(spu)
            else:
                missing_drafts.append(spu.spu_key)
                continue
        listings[spu.spu_key] = draft

    if missing_drafts and not args.mock_drafts:
        print(
            f"error: {len(missing_drafts)} 个 SPU 缺 ListingDraft（drafts/{{spu_key}}.json）；"
            f"请先跑 T-303 listing_generator，或加 --mock-drafts。\n"
            f"缺失：{missing_drafts[:5]}{'...' if len(missing_drafts) > 5 else ''}",
            file=sys.stderr,
        )
        return 2

    # Categories + attributes
    rules = load_category_rules()
    defaults = load_attr_defaults()
    cat_assign: Dict[str, str] = {}
    attrs_map: Dict[str, Dict[str, object]] = {}
    for spu in spus:
        listing = listings.get(spu.spu_key)
        cat_id = pick_category(spu, listing, rules=rules)
        cat_assign[spu.spu_key] = cat_id
        attrs_map[spu.spu_key] = fill_attributes(cat_id, spu, listing, defaults=defaults)

    # Images — by default skip network; users can extend by passing image_refs map.
    image_refs: Dict[str, Optional[ImageRef]] = {}
    if args.no_image_fetch:
        for spu in spus:
            for v in spu.variants:
                image_refs[v.jan] = None
    else:
        # Try cache only — never invoke network here (CLI is meant to be cheap).
        # Real network image lookups should be a separate batch step.
        try:
            from image_finder import ImageRefCache
            cache = ImageRefCache()
            for spu in spus:
                for v in spu.variants:
                    image_refs[v.jan] = cache.get(v.jan)
        except Exception as e:  # noqa: BLE001
            print(
                f"warning: 主图缓存读取失败 ({e})；所有主图按缺失处理",
                file=sys.stderr,
            )
            for spu in spus:
                for v in spu.variants:
                    image_refs[v.jan] = None

    results = export_to_xlsx(
        spus,
        listings,
        attrs_map,
        image_refs,
        out_dir,
        category_assignment=cat_assign,
        template_dir=template_dir,
    )

    for r in results:
        print(
            f"OK [{r.category_id}] {r.path}  "
            f"({r.spu_count} SPU, {r.row_count} variant rows, "
            f"{len(r.missing_image_jans)} missing images)"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
