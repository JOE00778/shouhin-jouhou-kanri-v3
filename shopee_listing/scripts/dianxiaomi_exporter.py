#!/usr/bin/env python3
"""店小秘批量上架 XLSX 导出器（T-310）

把 T-301~T-305 的输出（SPU + ListingDraft + 类目 + 属性 + 主图）组装成店小秘
后台能直接吃的 34 列 XLSX 文件。**单个 xlsx**（不像 T-306 按类目分文件）。

模板：`shopee-listing/店小秘上架模板/template_shopee_global.xlsx`（仅 1 行表头）。
该 xlsx 含 activePane="" 缺陷，openpyxl 直接 load_workbook 会抛 ValueError，
故沿用 T-306 方案 A：calamine 读模板表头 → openpyxl 写新文件。

34 列结构（详见 docs/03-dianxiaomi-template-fields.md）：
    [0]  分类ID            ← T-304 pick_category
    [1]  产品属性          ← T-304 fill_attributes 转 JSON
    [2]  Parent SKU        ← SPU.spu_key
    [3]  产品标题          ← ListingDraft.title           ✅ 必填
    [4]  产品描述          ← ListingDraft.description     ✅ 必填
    [5]  SKU              ← variant.jan（单品留空）
    [6]  变种名称          ← variant_attrs values 逗号合并
    [7]  变种属性名称一    ← variant_attrs.keys()[0]
    [8]  变种属性名称二    ← variant_attrs.keys()[1]
    [9]  变种属性值一      ← variant_attrs.values()[0]
    [10] 变种属性值二      ← variant_attrs.values()[1]
    [11] 价格              ← 留空（运营填）
    [12] 库存              ← 留空（运营填）
    [13] 重量(kg)          ← variant.weight_grams / 1000
    [14] 主图URL           ← ImageRef.url
    [15-22] 附图1-8        ← 留空
    [23] 变种图            ← 留空
    [24-26] 长宽高         ← 留空
    [27] 产品保存状况      ← "全新"
    [28] 发货期            ← 留空
    [29] 来源URL           ← 留空
    [30] 尺码图            ← 留空
    [31] 备注              ← 留空
    [32] 品牌              ← {"brand_id":0,"original_brand_name":<...>}
    [33] 视频URL           ← 留空

多变种合并：同 Parent SKU 多 variants 写多行；每行 Parent SKU/标题/描述
一致（参考 example.xlsx 第 2 行的写法）。

Module API:
    from dianxiaomi_exporter import export_to_dianxiaomi
    path = export_to_dianxiaomi(spus, listings, attributes, image_refs, out_path)

CLI:
    python3 scripts/dianxiaomi_exporter.py \\
        --spu-input examples/simple_spu_sample.csv \\
        --drafts-dir drafts/ \\
        --out-dir exports/ \\
        [--mock-drafts]
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence

import python_calamine
from openpyxl import Workbook

# ---- sibling imports -------------------------------------------------------- #
_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from parse_spu_input import SPU, Variant  # noqa: E402
from parse_simple_spu import parse_simple_spu_csv  # noqa: E402
from listing_generator import ListingDraft  # noqa: E402
from category_mapper import (  # noqa: E402
    fill_attributes,
    load_attr_defaults,
    load_category_rules,
    pick_category,
)
from image_finder import ImageRef  # noqa: E402

PROJECT_ROOT = _SCRIPTS_DIR.parent
DEFAULT_TEMPLATE_DIR = PROJECT_ROOT / "店小秘上架模板"
DEFAULT_TEMPLATE_FILE = "template_shopee_global.xlsx"
DEFAULT_DRAFTS_DIR = PROJECT_ROOT / "drafts"

OUTPUT_SHEET_NAME = "Sheet1"
EXPECTED_COLUMNS = 34
DEFAULT_CONDITION = "全新"  # [27]

# 列索引常量（0-based 对应模板列；写 xlsx 时要 +1）
COL_CATEGORY_ID = 0
COL_PRODUCT_ATTRS = 1
COL_PARENT_SKU = 2
COL_TITLE = 3
COL_DESCRIPTION = 4
COL_SKU = 5
COL_VARIANT_NAME = 6
COL_VARIANT_ATTR_1_NAME = 7
COL_VARIANT_ATTR_2_NAME = 8
COL_VARIANT_ATTR_1_VALUE = 9
COL_VARIANT_ATTR_2_VALUE = 10
COL_PRICE = 11
COL_STOCK = 12
COL_WEIGHT_KG = 13
COL_MAIN_IMAGE_URL = 14
# 15-22 附图1-8
# 23 变种图
# 24-26 长/宽/高
COL_CONDITION = 27
# 28 发货期
# 29 来源URL
# 30 尺码图
# 31 备注
COL_BRAND = 32
# 33 视频URL


# --------------------------------------------------------------------------- #
# Errors
# --------------------------------------------------------------------------- #


class DianxiaomiExporterError(Exception):
    """Raised on configuration / IO failures."""


# --------------------------------------------------------------------------- #
# Template header loader
# --------------------------------------------------------------------------- #


def load_template_header(
    *,
    template_dir: Path = DEFAULT_TEMPLATE_DIR,
    template_file: str = DEFAULT_TEMPLATE_FILE,
) -> List[str]:
    """读店小秘空模板的表头（仅 1 行）。"""
    path = Path(template_dir) / template_file
    if not path.exists():
        raise DianxiaomiExporterError(f"模板文件不存在: {path}")

    wb = python_calamine.CalamineWorkbook.from_path(str(path))
    if not wb.sheet_names:
        raise DianxiaomiExporterError(f"模板无 sheet: {path.name}")
    sheet = wb.get_sheet_by_name(wb.sheet_names[0])
    rows = sheet.to_python()
    if not rows:
        raise DianxiaomiExporterError(f"模板首 sheet 为空: {path.name}")

    header = ["" if c is None else str(c) for c in rows[0]]
    if len(header) != EXPECTED_COLUMNS:
        raise DianxiaomiExporterError(
            f"模板表头列数 {len(header)} != {EXPECTED_COLUMNS}（{path.name}）"
        )
    return header


# --------------------------------------------------------------------------- #
# Field builders
# --------------------------------------------------------------------------- #


def _format_brand_json(brand_normalized: Optional[str]) -> str:
    """品牌 JSON 串：`{"brand_id":0,"original_brand_name":"<brand>"}`。

    缺值时填 "NoBrand"（与示例文件一致）。
    """
    name = (brand_normalized or "").strip() or "NoBrand"
    return json.dumps(
        {"brand_id": 0, "original_brand_name": name},
        ensure_ascii=False,
    )


def _format_attributes_json(attributes: Mapping[str, Any]) -> str:
    """T-304 attributes dict → 店小秘 JSON 数组。

    输入形如 `{"100037": "Japan", "100036": "Solid"}`。
    输出形如：
      [{"attribute_id":"100037","original_attribute_name":"100037",
        "attribute_value_list":[{"value_id":"","original_value_name":"Japan"}]}, ...]

    属性名暂用 attribute_id（T-304 不返回名称；待 Boss 给细分类目 attr 字典再扩）。
    若属性 dict 为空，返回空字符串（让运营在店小秘后台填）。
    """
    if not attributes:
        return ""

    items = []
    for attr_id, raw_value in attributes.items():
        if raw_value is None:
            continue
        # 容许 list 多值
        if isinstance(raw_value, (list, tuple)):
            values = [str(v).strip() for v in raw_value if str(v).strip()]
        else:
            sv = str(raw_value).strip()
            values = [sv] if sv else []
        if not values:
            continue
        items.append(
            {
                "attribute_id": str(attr_id),
                "original_attribute_name": str(attr_id),
                "attribute_value_list": [
                    {"value_id": "", "original_value_name": v} for v in values
                ],
            }
        )

    if not items:
        return ""
    return json.dumps(items, ensure_ascii=False)


def _variant_attr_pairs(variant: Variant) -> List[tuple[str, str]]:
    """Return ordered (name, value) pairs from variant_attrs, skipping empties."""
    out: List[tuple[str, str]] = []
    for k, v in variant.variant_attrs.items():
        if v is None:
            continue
        sv = str(v).strip()
        if not sv:
            continue
        name = (k or "").strip()
        if not name:
            continue
        out.append((name, sv))
    return out


def _make_row(
    *,
    width: int,
    spu: SPU,
    variant: Variant,
    listing: ListingDraft,
    category_id: str,
    attributes: Mapping[str, Any],
    image_ref: Optional[ImageRef],
    is_multi_variant: bool,
) -> List[Any]:
    """Build a 34-column data row.

    Per docs/03 + example.xlsx row 2 convention: title/description/Parent SKU
    are filled identically on **every** variant row of the same SPU.

    `is_multi_variant`: True 当 SPU 有 >=2 个 variants（按例子 row 3 的写法，
    单品 SKU/变种名称留空）。
    """
    row: List[Any] = ["" for _ in range(width)]

    pairs = _variant_attr_pairs(variant)

    # [0] 分类 ID
    row[COL_CATEGORY_ID] = category_id or ""
    # [1] 产品属性 JSON
    row[COL_PRODUCT_ATTRS] = _format_attributes_json(attributes)
    # [2] Parent SKU
    row[COL_PARENT_SKU] = spu.spu_key
    # [3] 标题
    row[COL_TITLE] = listing.title
    # [4] 描述
    row[COL_DESCRIPTION] = listing.description

    # [5] SKU + [6-10] 变种字段：单品 SPU 留空，多变种 SPU 填
    if is_multi_variant:
        row[COL_SKU] = variant.jan
        # 变种名称：用 variant_attrs.values() 逗号合并；缺则用 jan 兜底（每行不可重复）
        if pairs:
            row[COL_VARIANT_NAME] = ",".join(p[1] for p in pairs)
            row[COL_VARIANT_ATTR_1_NAME] = pairs[0][0]
            row[COL_VARIANT_ATTR_1_VALUE] = pairs[0][1]
            if len(pairs) >= 2:
                row[COL_VARIANT_ATTR_2_NAME] = pairs[1][0]
                row[COL_VARIANT_ATTR_2_VALUE] = pairs[1][1]
        else:
            # 多 variant 但没有 attrs → 用 jan 当变种名称（保证唯一）
            row[COL_VARIANT_NAME] = variant.jan
    # else 单品：SKU/变种名称等留空（已是 ""）

    # [11] 价格 [12] 库存 — 留空
    # [13] 重量(kg)
    if variant.weight_grams is not None:
        try:
            row[COL_WEIGHT_KG] = round(float(variant.weight_grams) / 1000.0, 4)
        except (TypeError, ValueError):
            row[COL_WEIGHT_KG] = ""

    # [14] 主图 URL
    if image_ref is not None and image_ref.url:
        row[COL_MAIN_IMAGE_URL] = image_ref.url

    # [27] 产品保存状况
    row[COL_CONDITION] = DEFAULT_CONDITION

    # [32] 品牌 JSON
    row[COL_BRAND] = _format_brand_json(listing.brand_normalized)

    return row


# --------------------------------------------------------------------------- #
# Workbook writer
# --------------------------------------------------------------------------- #


def _write_workbook(
    *,
    header: Sequence[str],
    data_rows: Sequence[Sequence[Any]],
    out_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET_NAME

    # Row 1: header verbatim
    for c_idx, val in enumerate(header, start=1):
        ws.cell(row=1, column=c_idx, value=val)

    # Data rows from row 2
    for d_idx, drow in enumerate(data_rows, start=2):
        for c_idx, val in enumerate(drow, start=1):
            ws.cell(row=d_idx, column=c_idx, value=val)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Public entry: export_to_dianxiaomi
# --------------------------------------------------------------------------- #


@dataclass
class DianxiaomiExportResult:
    path: Path
    spu_count: int
    row_count: int
    missing_image_jans: List[str]


def export_to_dianxiaomi(
    spus: Sequence[SPU],
    listings: Mapping[str, ListingDraft],
    attributes: Mapping[str, Mapping[str, Any]],
    image_refs: Mapping[str, Optional[ImageRef]],
    out_path: Path,
    *,
    category_assignment: Optional[Mapping[str, str]] = None,
    template_dir: Path = DEFAULT_TEMPLATE_DIR,
    template_file: str = DEFAULT_TEMPLATE_FILE,
    warn_logger=None,
) -> DianxiaomiExportResult:
    """单 xlsx 导出（所有 SPU 写入同一文件）。

    Args:
        spus: T-301/T-310 的 SPU 列表
        listings: spu_key → ListingDraft（T-303）
        attributes: spu_key → {attr_id: value}（T-304）
        image_refs: jan → ImageRef|None（T-305）
        out_path: 输出 xlsx 完整路径
        category_assignment: spu_key → 分类 ID（None 时调 pick_category）
        template_dir / template_file: 店小秘空模板位置
        warn_logger: 缺图警告 sink，None 时打 stderr

    Returns:
        DianxiaomiExportResult
    """
    if warn_logger is None:
        warn_logger = lambda msg: print(f"warning: {msg}", file=sys.stderr)  # noqa: E731

    header = load_template_header(
        template_dir=template_dir,
        template_file=template_file,
    )
    width = len(header)

    rows: List[List[Any]] = []
    missing_imgs: List[str] = []

    for spu in spus:
        listing = listings.get(spu.spu_key)
        if listing is None:
            warn_logger(
                f"SPU {spu.spu_key} 没有 ListingDraft，跳过；请先跑 T-303"
            )
            continue

        if category_assignment and spu.spu_key in category_assignment:
            cat_id = category_assignment[spu.spu_key]
        else:
            try:
                cat_id = pick_category(spu, listing)
            except Exception as e:  # noqa: BLE001
                warn_logger(
                    f"SPU {spu.spu_key} pick_category 失败 ({e})，分类 ID 留空"
                )
                cat_id = ""

        attrs = attributes.get(spu.spu_key, {}) or {}

        if not spu.variants:
            # 没有 variant 的 SPU（理论上不应该出现），跳过
            warn_logger(f"SPU {spu.spu_key} 无 variant，跳过")
            continue

        is_multi = len(spu.variants) >= 2
        for variant in spu.variants:
            img_ref = image_refs.get(variant.jan)
            if img_ref is None:
                missing_imgs.append(variant.jan)
                warn_logger(
                    f"SPU {spu.spu_key} variant {variant.jan} 缺主图（导出时留空）"
                )
            row = _make_row(
                width=width,
                spu=spu,
                variant=variant,
                listing=listing,
                category_id=str(cat_id) if cat_id else "",
                attributes=attrs,
                image_ref=img_ref,
                is_multi_variant=is_multi,
            )
            rows.append(row)

    out_path = Path(out_path)
    _write_workbook(header=header, data_rows=rows, out_path=out_path)

    return DianxiaomiExportResult(
        path=out_path,
        spu_count=len([s for s in spus if s.spu_key in listings]),
        row_count=len(rows),
        missing_image_jans=missing_imgs,
    )


# --------------------------------------------------------------------------- #
# CLI helpers — load drafts / images
# --------------------------------------------------------------------------- #


def _load_draft_from_disk(drafts_dir: Path, spu_key: str) -> Optional[ListingDraft]:
    p = drafts_dir / f"{spu_key}.json"
    if not p.exists():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        print(f"warning: 解析 {p} 失败: {e}", file=sys.stderr)
        return None
    return ListingDraft(
        title=data.get("title", ""),
        description=data.get("description", ""),
        key_features=list(data.get("key_features", []) or []),
        how_to_use=list(data.get("how_to_use", []) or []),
        ingredients=data.get("ingredients"),
        spec_json=dict(data.get("spec_json", {}) or {}),
        brand_normalized=data.get("brand_normalized", ""),
        hook=data.get("hook", ""),
        model=data.get("model", ""),
        spu_key=data.get("spu_key", spu_key),
    )


def _make_mock_draft(spu: SPU) -> ListingDraft:
    return ListingDraft(
        title=f"<MOCK> {spu.spu_key} — Direct from Japan",
        description=(
            f"<MOCK description for {spu.spu_key}>\n"
            "[NOTE] Placeholder generated by dianxiaomi_exporter --mock-drafts. "
            "Replace by running listing_generator (T-303) with ANTHROPIC_API_KEY set. "
            "Smikie Japan will deliver."
        ),
        key_features=["<mock feature 1>", "<mock feature 2>"],
        how_to_use=["<mock step 1>"],
        ingredients=None,
        spec_json={},
        brand_normalized="",
        hook="Direct from Japan",
        model="mock",
        spu_key=spu.spu_key,
    )


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #


def _build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="dianxiaomi_exporter",
        description="店小秘批量上架 XLSX 导出（34 列单文件，T-310）",
    )
    p.add_argument(
        "--spu-input",
        required=True,
        help="极简 SPU CSV 路径（A=SPU, B=SKU 两列）",
    )
    p.add_argument(
        "--drafts-dir",
        default=str(DEFAULT_DRAFTS_DIR),
        help=f"ListingDraft JSON 目录（默认 {DEFAULT_DRAFTS_DIR}）",
    )
    p.add_argument("--out-dir", required=True, help="XLSX 输出目录")
    p.add_argument(
        "--mock-drafts",
        action="store_true",
        help="drafts/<spu_key>.json 缺失时用 <MOCK> 占位（dry-run 用）",
    )
    p.add_argument(
        "--template-dir",
        default=str(DEFAULT_TEMPLATE_DIR),
        help=f"店小秘模板目录（默认 {DEFAULT_TEMPLATE_DIR}）",
    )
    p.add_argument(
        "--template-file",
        default=DEFAULT_TEMPLATE_FILE,
        help=f"店小秘模板文件名（默认 {DEFAULT_TEMPLATE_FILE}）",
    )
    return p


def main(argv: Optional[List[str]] = None) -> int:
    args = _build_argparser().parse_args(argv)

    spu_path = Path(args.spu_input)
    drafts_dir = Path(args.drafts_dir)
    out_dir = Path(args.out_dir)
    template_dir = Path(args.template_dir)

    try:
        spus = parse_simple_spu_csv(spu_path)
    except Exception as e:  # noqa: BLE001
        print(f"error: 解析 SPU CSV 失败: {e}", file=sys.stderr)
        return 2
    if not spus:
        print("error: SPU 输入为空", file=sys.stderr)
        return 2

    # Drafts
    listings: Dict[str, ListingDraft] = {}
    missing_drafts: List[str] = []
    for spu in spus:
        draft = _load_draft_from_disk(drafts_dir, spu.spu_key)
        if draft is None:
            if args.mock_drafts:
                draft = _make_mock_draft(spu)
            else:
                missing_drafts.append(spu.spu_key)
                continue
        listings[spu.spu_key] = draft

    if missing_drafts and not args.mock_drafts:
        print(
            f"error: {len(missing_drafts)} 个 SPU 缺 ListingDraft（drafts/{{spu_key}}.json）；"
            f"请先跑 T-303 listing_generator，或加 --mock-drafts。\n"
            f"缺失：{missing_drafts[:5]}{'...' if len(missing_drafts) > 5 else ''}",
            file=sys.stderr,
        )
        return 2

    # Categories + attributes
    try:
        rules = load_category_rules()
        defaults = load_attr_defaults()
    except Exception as e:  # noqa: BLE001
        print(
            f"warning: 类目规则加载失败 ({e})；分类/属性留空",
            file=sys.stderr,
        )
        rules = None
        defaults = None

    cat_assign: Dict[str, str] = {}
    attrs_map: Dict[str, Dict[str, Any]] = {}
    for spu in spus:
        listing = listings.get(spu.spu_key)
        if rules is not None:
            try:
                cat_id = pick_category(spu, listing, rules=rules)
            except Exception:
                cat_id = ""
        else:
            cat_id = ""
        cat_assign[spu.spu_key] = cat_id
        if defaults is not None and cat_id:
            try:
                attrs_map[spu.spu_key] = fill_attributes(
                    cat_id, spu, listing, defaults=defaults
                )
            except Exception:
                attrs_map[spu.spu_key] = {}
        else:
            attrs_map[spu.spu_key] = {}

    # Images — try cache only (CLI 不主动联网)
    image_refs: Dict[str, Optional[ImageRef]] = {}
    try:
        from image_finder import ImageRefCache

        cache = ImageRefCache()
        for spu in spus:
            for v in spu.variants:
                image_refs[v.jan] = cache.get(v.jan)
    except Exception as e:  # noqa: BLE001
        print(
            f"warning: 主图缓存读取失败 ({e})；所有主图按缺失处理",
            file=sys.stderr,
        )
        for spu in spus:
            for v in spu.variants:
                image_refs[v.jan] = None

    # Output path: out_dir/dianxiaomi_shopee_global_{ts}.xlsx
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"dianxiaomi_shopee_global_{timestamp}.xlsx"

    result = export_to_dianxiaomi(
        spus,
        listings,
        attrs_map,
        image_refs,
        out_path,
        category_assignment=cat_assign,
        template_dir=template_dir,
        template_file=args.template_file,
    )

    print(
        f"OK {result.path}  "
        f"({result.spu_count} SPU, {result.row_count} variant rows, "
        f"{len(result.missing_image_jans)} missing images)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
