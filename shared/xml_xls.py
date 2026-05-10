"""NetSuite SpreadsheetML XML 解析器 + OOXML (.xlsx) fallback。

NetSuite Item Search / Report 默认导出 SpreadsheetML 2003 XML（扩展名 .xls 但
内部是 XML，pandas/xlrd/openpyxl 都不认）。如果 Boss 在 NetSuite 选了 'Excel
(.xlsx)' 而不是 'XML'，文件就是真 OOXML zip — 走 openpyxl fallback。

使用：
    from shared.xml_xls import parse_smart   # 推荐: 自动按文件头 fallback
    rows = parse_smart(path, header_row=6)

    # 旧 API 仍保留, 仅认 SpreadsheetML XML:
    from shared.xml_xls import iter_rows, parse_to_dicts
"""
from __future__ import annotations

import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Iterator

NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
NS_INDEX_ATTR = "{urn:schemas-microsoft-com:office:spreadsheet}Index"


def iter_rows(path: str | Path, sheet_index: int = 0) -> Iterator[list[str | None]]:
    """逐行迭代 SpreadsheetML 文件。

    每行返回 list[str | None]，单元格按 ss:Index 属性正确对齐
    （NetSuite 偶尔会跳过空单元格用 ss:Index 标位置）。
    """
    tree = ET.parse(path)
    root = tree.getroot()
    worksheets = root.findall("ss:Worksheet", NS)
    if sheet_index >= len(worksheets):
        raise IndexError(f"Sheet index {sheet_index} out of range (have {len(worksheets)})")
    ws = worksheets[sheet_index]
    table = ws.find("ss:Table", NS)
    if table is None:
        return

    for row in table.findall("ss:Row", NS):
        cells: list[str | None] = []
        col_idx = 1  # SpreadsheetML 列索引从 1 开始
        for cell in row.findall("ss:Cell", NS):
            # 处理 ss:Index 跳跃（被跳过的列填 None）
            cell_idx_attr = cell.attrib.get(NS_INDEX_ATTR)
            if cell_idx_attr:
                target = int(cell_idx_attr)
                while col_idx < target:
                    cells.append(None)
                    col_idx += 1
            data = cell.find("ss:Data", NS)
            cells.append(data.text if data is not None else None)
            col_idx += 1
        yield cells


def parse_to_dicts(
    path: str | Path,
    *,
    header_row: int = 0,
    sheet_index: int = 0,
    skip_empty_rows: bool = True,
) -> list[dict[str, str | None]]:
    """读 SpreadsheetML 文件 → 返回 list[dict]。

    Args:
        path: 文件路径
        header_row: 表头在第几行（0-indexed）。NetSuite Saved Search 导出 = 0；
                   NetSuite Report 导出 = 6（前 6 行是公司名 / 标题 / 期间 / 空行）
        sheet_index: 工作表索引（默认第 0 个）
        skip_empty_rows: 跳过完全空白的行
    """
    rows_iter = list(iter_rows(path, sheet_index=sheet_index))
    if header_row >= len(rows_iter):
        return []

    headers = rows_iter[header_row]
    # 清理 None 列名（替换为占位符避免 dict key 冲突）
    headers = [
        (h.strip() if h else f"_col{i}") for i, h in enumerate(headers)
    ]

    out: list[dict[str, str | None]] = []
    for row in rows_iter[header_row + 1 :]:
        if skip_empty_rows and all(v is None or v == "" for v in row):
            continue
        # 用 zip 对齐，多余列丢弃，缺少列填 None
        d = dict(zip(headers, row))
        # 补齐：确保所有 header 都有 key
        for h in headers:
            d.setdefault(h, None)
        out.append(d)
    return out


def _sniff_format(path: str | Path) -> str:
    """按前几个字节判断文件实际格式。返回 'xml' / 'ooxml' / 'ole2' / 'unknown'."""
    with open(path, "rb") as f:
        head = f.read(80)
    if head[:2] == b"PK":
        return "ooxml"
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "ole2"
    if b"<?xml" in head:
        return "xml"
    return "unknown"


def _iter_rows_ooxml(path: str | Path, sheet_index: int = 0) -> Iterator[list]:
    """用 openpyxl 读真 .xlsx (OOXML zip)。"""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        sheet_names = wb.sheetnames
        if sheet_index >= len(sheet_names):
            raise IndexError(
                f"Sheet index {sheet_index} out of range (have {len(sheet_names)})"
            )
        ws = wb[sheet_names[sheet_index]]
        for row in ws.iter_rows(values_only=True):
            yield [
                (str(v).strip() if isinstance(v, str) else v) for v in row
            ]
    finally:
        wb.close()


def iter_rows_smart(path: str | Path, sheet_index: int = 0) -> Iterator[list]:
    """统一入口: 按文件头自动选 SpreadsheetML XML / OOXML 解析。

    OLE2 (真二进制 .xls) 暂不支持 — NetSuite 不会导出这种格式。
    """
    fmt = _sniff_format(path)
    if fmt == "ooxml":
        yield from _iter_rows_ooxml(path, sheet_index=sheet_index)
    elif fmt in ("xml", "unknown"):
        # 'unknown' 时也试 XML — 保持向后兼容（旧测试 fixture 可能没 BOM/preamble）
        yield from iter_rows(path, sheet_index=sheet_index)
    else:  # ole2
        raise RuntimeError(
            f"文件是 .xls (OLE2 binary) 真二进制格式, 当前不支持。"
            f"请在 NetSuite 导出时选 'XML' 或 'Excel (.xlsx)' 格式。文件: {path}"
        )


def parse_smart(
    path: str | Path,
    *,
    header_row: int = 0,
    sheet_index: int = 0,
    skip_empty_rows: bool = True,
) -> list[dict]:
    """parse_to_dicts 的多格式版本: SpreadsheetML XML + OOXML 自动 fallback。"""
    rows_iter = list(iter_rows_smart(path, sheet_index=sheet_index))
    if header_row >= len(rows_iter):
        return []

    headers = rows_iter[header_row]
    headers = [
        ((h.strip() if isinstance(h, str) else str(h)) if h not in (None, "") else f"_col{i}")
        for i, h in enumerate(headers)
    ]

    out: list[dict] = []
    for row in rows_iter[header_row + 1:]:
        if skip_empty_rows and all(v is None or v == "" for v in row):
            continue
        d = dict(zip(headers, row))
        for h in headers:
            d.setdefault(h, None)
        out.append(d)
    return out


def detect_header_row(path: str | Path, *, max_check: int = 15) -> int:
    """启发式检测 NetSuite 导出的表头行。

    NetSuite 报表 preamble 特征：
    - 前几行只有 1 个非空单元格（公司名 / 标题 / 期间）
    - 表头行通常 ≥ 4 个非空列

    返回检测到的 header 行号；如果检测不到合理候选，返回 0。
    """
    rows_iter = []
    for i, row in enumerate(iter_rows_smart(path)):
        rows_iter.append(row)
        if i >= max_check:
            break

    for i, row in enumerate(rows_iter):
        non_empty = sum(1 for v in row if v not in (None, ""))
        if non_empty >= 4:
            return i
    return 0
