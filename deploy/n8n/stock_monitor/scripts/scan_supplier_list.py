"""月度对比 仕入先管理リスト JAN 列差异。

读取 /Users/joe/CC/商品信息管理/仕入先管理リスト.xlsx 中各供应商 sheet，
与上月快照对比，输出 alerts list（新增/删除 JAN）。
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

from config import STATE_DIR

SUPPLIER_LIST_PATH = Path("/Users/joe/CC/商品信息管理/仕入先管理リスト.xlsx")
SUPPLIER_SNAPSHOT_FILE = STATE_DIR / "supplier_jan_snapshot.json"


def load_supplier_sheet(sheet_name: str) -> set[str]:
    """从指定 sheet 读取 JAN 列（去重）。返回 JAN 集合。"""
    if openpyxl is None:
        print(f"Warning: openpyxl not installed, skipping {sheet_name}")
        return set()

    if not SUPPLIER_LIST_PATH.exists():
        print(f"Warning: {SUPPLIER_LIST_PATH} not found")
        return set()

    try:
        wb = openpyxl.load_workbook(str(SUPPLIER_LIST_PATH), data_only=True)
        if sheet_name not in wb.sheetnames:
            return set()

        ws = wb[sheet_name]
        jans = set()
        # 假设 JAN 在 A 列（或通过标题行查找）
        for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过标题行
            if row and row[0]:
                jan = str(row[0]).strip()
                if jan and jan != "None":
                    jans.add(jan)
        return jans
    except Exception as e:
        print(f"Error loading sheet {sheet_name}: {e}")
        return set()


def load_all_suppliers() -> dict[str, set[str]]:
    """读取所有供应商 sheet 的 JAN 列。返回 {supplier_name: set(JAN)}。"""
    supplier_names = [
        "NEW WIND", "中央物産", "菅野", "Maple",
        "NEW WIND-1", "新风-2", "供应商1", "供应商2"  # 备选
    ]

    result = {}
    for name in supplier_names:
        jans = load_supplier_sheet(name)
        if jans:
            result[name] = jans
    return result


def load_prev_snapshot() -> dict[str, set[str]]:
    """加载上月快照。返回 {supplier_name: set(JAN)}。"""
    if SUPPLIER_SNAPSHOT_FILE.exists():
        data = json.loads(SUPPLIER_SNAPSHOT_FILE.read_text(encoding="utf-8"))
        return {k: set(v) for k, v in data.items()}
    return {}


def save_snapshot(suppliers: dict[str, set[str]]):
    """保存当月快照。"""
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    data = {k: sorted(v) for k, v in suppliers.items()}
    SUPPLIER_SNAPSHOT_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def diff_suppliers(prev: dict[str, set[str]], curr: dict[str, set[str]]) -> list[dict]:
    """
    比较前后快照，返回 alerts list。
    alerts 字段：jan / supplier / signal_type ('NEW' 或 '削除')
    """
    alerts = []

    # 遍历当前所有供应商
    for supplier, curr_jans in curr.items():
        prev_jans = prev.get(supplier, set())

        # 新增 JAN
        new_jans = curr_jans - prev_jans
        for jan in new_jans:
            alerts.append({
                "jan": jan,
                "supplier": supplier,
                "signal_type": "NEW",
                "source": "supplier_list",
                "detected_at": datetime.now().isoformat(timespec="seconds"),
            })

        # 删除 JAN
        deleted_jans = prev_jans - curr_jans
        for jan in deleted_jans:
            alerts.append({
                "jan": jan,
                "supplier": supplier,
                "signal_type": "削除",
                "source": "supplier_list",
                "detected_at": datetime.now().isoformat(timespec="seconds"),
            })

    return alerts


def main():
    curr = load_all_suppliers()
    if not curr:
        print("No supplier sheets found or openpyxl not installed")
        return

    prev = load_prev_snapshot()
    alerts = diff_suppliers(prev, curr)

    # 保存本月快照
    save_snapshot(curr)

    # 输出 alerts
    if alerts:
        print(f"Found {len(alerts)} alerts:")
        for alert in alerts:
            print(f"  {alert['supplier']}: {alert['jan']} ({alert['signal_type']})")
    else:
        print("No supplier list changes detected")

    return alerts


if __name__ == "__main__":
    main()
