"""Shopee 拨款数据导入（ph.mtkshop.ph.income.xlsx）。

从 4 个 sheet 导入：
- Income → shopee_payouts（拨款主档）
- Service Fee Details → shopee_fees（服务费明细，需要枚举化 fee_type）
- Adjustment → shopee_adjustments（调整记录）
- Summary → 跳过（汇总数据）
"""
from __future__ import annotations

import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


def _to_float(value) -> float | None:
    """尝试转为 float，失败返回 None。"""
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def ingest_income(ws, conn: sqlite3.Connection) -> int:
    """Income sheet → shopee_payouts。

    header row：卖家帐号 / 付款ID / 收款渠道 / 拨款时间 / ...其他
    data rows：逐行展开为 payout 记录
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    inserted = 0

    for row in rows:
        if not row or all(v is None for v in row):
            continue

        seller_account = str(row[0] or "").strip() if row and len(row) > 0 and row[0] is not None else None
        payout_id = str(row[1] or "").strip() if row and len(row) > 1 and row[1] is not None else None
        channel = str(row[2] or "").strip() if row and len(row) > 2 and row[2] is not None else None
        payout_date_raw = row[3] if row and len(row) > 3 else None

        # 拨款时间格式解析（可能是日期或字符串）
        payout_date = None
        if payout_date_raw:
            if hasattr(payout_date_raw, "date"):  # datetime object
                payout_date = str(payout_date_raw.date())
            else:
                payout_date = str(payout_date_raw).strip()

        # 找到金额列（不在前 4 列，需要从其他位置估计）
        # 由于 Income sheet 结构不完全明确，暂时用 None
        total_payout = None
        currency = None

        if seller_account and payout_date:
            try:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO shopee_payouts
                    (payout_id, seller_account, channel, payout_date, total_payout, currency)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (payout_id or seller_account, seller_account, channel, payout_date, total_payout, currency),
                )
                inserted += 1
            except sqlite3.Error as e:
                print(f"  Warning: 拨款记录写入失败：{e}", file=sys.stderr)

    return inserted


def ingest_service_fees(ws, conn: sqlite3.Connection) -> int:
    """Service Fee Details sheet → shopee_fees。

    header：编号 / 订单编号 / CB Infrastructure Fee / CB MDV 4% / CB MDV 5% / CB MP Platform Shipping Fee
    每行的多个费用列 → 多个 fee 记录
    """
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # row 1=title, row 2=header
    inserted = 0

    # 字段映射：列索引 → (fee_type, column_index)
    fee_columns = {
        2: "cb_infrastructure",
        3: "cb_mdv_4",
        4: "cb_mdv_5",
        5: "platform_shipping",
    }

    for row in rows:
        if not row or all(v is None for v in row):
            continue

        order_no = str(row[1] or "").strip() if row and len(row) > 1 and row[1] is not None else None

        if not order_no:
            continue

        # 遍历费用列
        for col_idx, fee_type in fee_columns.items():
            if row and len(row) > col_idx and row[col_idx] is not None:
                amount = _to_float(row[col_idx])
                if amount is not None:
                    try:
                        conn.execute(
                            """
                            INSERT INTO shopee_fees (order_no, fee_type, amount)
                            VALUES (?, ?, ?)
                            """,
                            (order_no, fee_type, amount),
                        )
                        inserted += 1
                    except sqlite3.Error as e:
                        print(f"  Warning: 费用记录写入失败 {order_no}/{fee_type}：{e}", file=sys.stderr)

    return inserted


def ingest_adjustments(ws, conn: sqlite3.Connection) -> int:
    """Adjustment sheet → shopee_adjustments。

    header：卖家帐号 / 付款ID / 收款渠道 / 拨款时间 / ...
    data rows：每行一条调整记录
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    inserted = 0

    for row in rows:
        if not row or all(v is None for v in row):
            continue

        seller_account = str(row[0] or "").strip() if row and len(row) > 0 and row[0] is not None else None
        payout_id = str(row[1] or "").strip() if row and len(row) > 1 and row[1] is not None else None
        # channel = str(row[2] or "").strip() if row and len(row) > 2 and row[2] is not None else None
        payout_date_raw = row[3] if row and len(row) > 3 else None

        # 拨款时间格式解析
        payout_date = None
        if payout_date_raw:
            if hasattr(payout_date_raw, "date"):  # datetime object
                payout_date = str(payout_date_raw.date())
            else:
                payout_date = str(payout_date_raw).strip()

        if seller_account:
            try:
                conn.execute(
                    """
                    INSERT INTO shopee_adjustments (seller_account, payout_id, payout_date)
                    VALUES (?, ?, ?)
                    """,
                    (seller_account, payout_id, payout_date),
                )
                inserted += 1
            except sqlite3.Error as e:
                print(f"  Warning: 调整记录写入失败：{e}", file=sys.stderr)

    return inserted


def main(excel_path: str, db_path: str = "data_warehouse/warehouse.db"):
    """主入口：读 Excel，执行 3 次 ingest，返回汇总。"""
    from data_warehouse.db.migrations import run as init_db

    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"错误：文件不存在 {excel_path}", file=sys.stderr)
        sys.exit(1)

    conn = init_db(db_path)

    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)

        print(f"[shopee_income] 开始导入 {excel_path.name}")
        print(f"  Sheets: {wb.sheetnames}")

        inserted = 0

        # Income → shopee_payouts
        if "Income" in wb.sheetnames:
            print(f"  处理 Income...")
            count = ingest_income(wb["Income"], conn)
            print(f"    ✓ shopee_payouts: {count} 条")
            inserted += count

        # Service Fee Details → shopee_fees
        if "Service Fee Details" in wb.sheetnames:
            print(f"  处理 Service Fee Details...")
            count = ingest_service_fees(wb["Service Fee Details"], conn)
            print(f"    ✓ shopee_fees: {count} 条")
            inserted += count

        # Adjustment → shopee_adjustments
        if "Adjustment" in wb.sheetnames:
            print(f"  处理 Adjustment...")
            count = ingest_adjustments(wb["Adjustment"], conn)
            print(f"    ✓ shopee_adjustments: {count} 条")
            inserted += count

        # Summary 跳过
        print(f"  跳过 Summary（汇总数据）")

        conn.commit()
        print(f"[shopee_income] 完成，共入库 {inserted} 条记录")

    finally:
        conn.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"用法：{sys.argv[0]} <excel_path> [db_path]", file=sys.stderr)
        sys.exit(1)

    excel_path = sys.argv[1]
    db_path = sys.argv[2] if len(sys.argv) > 2 else "data_warehouse/warehouse.db"
    main(excel_path, db_path)
