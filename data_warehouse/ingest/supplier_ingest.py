"""仕入先管理リスト.xlsx → supplier_quote テーブル ingester.

入力: 仕入先管理リスト.xlsx (35 sheet, うち 28〜29 が仕入先個別 sheet)
出力: supplier_quote (supplier_name × jan の報価, zone 分類付き)

仕入先個別 sheet の標準列:
  JAN | 商品名 | 単価 | ロット | (空) | (空) | 注文最低金額 | 発注条件 | 納期
変体:
  - NEW WIND: JAN 列なし (商品名のみ) → スキップ (関連不可)
  - 共和/大木: 「単価（JDへの送料込み）」列が前に追加 → 原始「単価」を採用
  - ハリマ: 商品名 が先頭, JAN は「JANコード」, ロットなし → 入数を lot_size に
  - 現金: 4 列のみ (JAN/商品名/単価/ロット)
"""
from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from shared.xml_xls import iter_rows_smart  # 兼容 xls/xlsx; 但 6.7MB xlsx は openpyxl 経由

import openpyxl


# 集計 sheet (仕入先個別ではない) — スキップ
_SUMMARY_SHEETS = {
    "仕入先-原価_AB類", "仕入先_原価_C類", "AB商品进货周期", "SUPABASE用",
    "FB_アイテム別売上12.4更新", "一元 Data12.4更新",
}

# sheet 名 → (zone, zone_rank, nst_supplier_code)  · Boss 2026-05-11 分類
# zone_rank: 1=JD直送 / 2=弁天経由 / 3=応急 / 4=前払い / 9=未分類
_SUPPLIER_ZONE: dict[str, tuple[str, int, str | None]] = {
    # --- JD 直送 (追加配送費なし) ---
    "NEW WIND": ("JD_DIRECT", 1, "0490 NEW WIND株式会社"),
    "中央物産": ("JD_DIRECT", 1, "0085 中央物産株式会社"),
    "菅野": ("JD_DIRECT", 1, "0376 菅野株式会社"),
    "Maple": ("JD_DIRECT", 1, "0486 Maple International株式会社"),
    "五洲": ("JD_DIRECT", 1, "0474 株式会社　五洲"),
    "アプライド": ("JD_DIRECT", 1, "0491 アプライド株式会社"),
    "王子国際": ("JD_DIRECT", 1, "0468 王子国際貿易株式会社"),
    "ハナモン": ("JD_DIRECT", 1, "0444 ハナモンワークス 合同会社"),
    "HK": ("JD_DIRECT", 1, "0073 株式会社　エィチ・ケイ"),
    "オンダ": ("JD_DIRECT", 1, "0025 株式会社オンダ"),
    "スケーター": ("JD_DIRECT", 1, "0479 スケーター株式会社"),
    "ファイン": ("JD_DIRECT", 1, "0258 株式会社 ファイン"),
    "新日配": ("JD_DIRECT", 1, "0469 株式会社 新日配薬品"),
    "グランジェ": ("JD_DIRECT", 1, "0256 株式会社 グランジェ"),
    "トラスコ中山": ("JD_DIRECT", 1, "0202 トラスコ中山株式会社"),
    "エンパイヤ自動車株式会社": ("JD_DIRECT", 1, "0020 エンパイヤ自動車株式会社（KONNGU'S）"),
    # --- 弁天倉庫経由 (中継費 +3%) ---
    "共和": ("BENTEN_TRANSIT", 2, "0077 大分共和株式会社"),
    "大木": ("BENTEN_TRANSIT", 2, "0197 大木化粧品株式会社"),
    # --- 応急 / 参考 ---
    "SD": ("EMERGENCY", 3, "0411 株式会社ラクーンコマース（スーパーデリバリー）"),
    "ハリマ": ("EMERGENCY", 3, "0402 ハリマ共和物産株式会社"),
    "カード仕入": ("EMERGENCY", 3, "0476 カード仕入れ"),
    # --- 前払い (現金) ---
    "流久": ("PREPAID", 4, "0435 株式会社 流久商事"),
    "冨森": ("PREPAID", 4, "0445 富森商事 株式会社"),
    "現金": ("PREPAID", 4, "0201 現金仕入れ"),
    "風雲商事": ("PREPAID", 4, "0482 風雲商事株式会社"),
    # --- 未分類 (Boss 未指定 — 後で校正) ---
    "カネイシ": ("OTHER", 9, "0457 カネイシ株式会社"),
    "京浜商事": ("OTHER", 9, "0504 京浜商事株式会社"),
    "若竹園": ("OTHER", 9, None),
    "森フォレスト": ("OTHER", 9, "0343 株式会社森フォレスト"),
    "エトワール海渡": ("OTHER", 9, None),
    "太田物産": ("OTHER", 9, "C000510 太田物産 株式会社"),
}


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _to_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", "").replace("円", "").strip()))
    except (ValueError, TypeError):
        return None


def _to_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _is_valid_jan(jan) -> bool:
    if jan is None:
        return False
    s = str(jan).strip()
    return s.isdigit() and 8 <= len(s) <= 14


def _map_columns(headers: list[str]) -> dict[str, int]:
    """sheet 表头 → {field: col_index} のマッピング。

    複数の「単価」列がある場合 (共和/大木): 「送料込み」を含まない方を採用。
    """
    m: dict[str, int] = {}
    price_candidates: list[int] = []
    for i, h in enumerate(headers):
        hs = str(h).strip() if h else ""
        if not hs:
            continue
        if "JAN" in hs.upper() and "jan" not in m:
            m["jan"] = i
        elif "商品名" in hs and "display_name" not in m:
            m["display_name"] = i
        elif hs == "単価" or (hs.startswith("単価") and "送料" not in hs and "込" not in hs):
            price_candidates.append(i)
        elif "送料込" in hs or "送料 込" in hs.replace("　", " "):
            # 共和/大木 の「単価（JDへの送料込み）」 — 採用しない (原始単価を使う)
            pass
        elif ("ケース入数" in hs or hs == "入数" or "入数" in hs) and "case_qty" not in m:
            m["case_qty"] = i
        elif "ロット" in hs and "lot_size" not in m:
            m["lot_size"] = i
        elif "注文最低金額" in hs and "min_order_amount" not in m:
            m["min_order_amount"] = i
        elif "発注条件" in hs and "order_condition" not in m:
            m["order_condition"] = i
        elif "納期" in hs and "lead_time_text" not in m:
            m["lead_time_text"] = i
    if price_candidates:
        m["unit_price"] = price_candidates[0]
    return m


def ingest_supplier_master(path, conn: sqlite3.Connection) -> dict:
    """仕入先管理リスト.xlsx を解析して supplier_quote を全量再構築。

    戻り値: {sheets_processed, rows_inserted, skipped_no_jan, sheets_skipped, warnings}
    """
    path = Path(path)
    # read_only=True だと一部 xlsx で dimension メタデータが壊れて 0 行になる → 通常モード
    wb = openpyxl.load_workbook(str(path), data_only=True)
    now = _now_iso()

    # 全量上書き
    conn.execute("DELETE FROM supplier_quote")

    sql = """
        INSERT OR REPLACE INTO supplier_quote (
            supplier_name, jan, display_name, unit_price, lot_size, case_qty,
            min_order_amount, order_condition, lead_time_text,
            zone, zone_rank, nst_supplier_code, source_sheet, imported_at
        ) VALUES (
            :supplier_name, :jan, :display_name, :unit_price, :lot_size, :case_qty,
            :min_order_amount, :order_condition, :lead_time_text,
            :zone, :zone_rank, :nst_supplier_code, :source_sheet, :imported_at
        )
    """

    sheets_done = rows_ins = skipped_no_jan = 0
    sheets_skipped: list[str] = []
    warnings: list[str] = []

    for name in wb.sheetnames:
        if name in _SUMMARY_SHEETS:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            sheets_skipped.append(f"{name} (空)")
            continue
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        colmap = _map_columns(headers)
        if "jan" not in colmap:
            sheets_skipped.append(f"{name} (JAN列なし)")
            warnings.append(f"{name}: JAN 列が見つからずスキップ (列: {[h for h in headers if h]})")
            continue
        zone, zone_rank, nst_code = _SUPPLIER_ZONE.get(name, ("OTHER", 9, None))
        if name not in _SUPPLIER_ZONE:
            warnings.append(f"{name}: zone 未分類 (OTHER) — 要校正")

        n_sheet = 0
        seen_jans: set[str] = set()
        for raw in rows[1:]:
            def cell(field):
                idx = colmap.get(field)
                if idx is None or idx >= len(raw):
                    return None
                return raw[idx]

            jan = _to_str(cell("jan"))
            if not _is_valid_jan(jan):
                skipped_no_jan += 1
                continue
            if jan in seen_jans:
                continue  # 同一 sheet 内重複は最初の行を採用
            seen_jans.add(jan)

            case_qty = _to_int(cell("case_qty"))
            lot_size = _to_int(cell("lot_size"))
            # ハリマ等 ロットなし → 入数を lot に
            if lot_size is None and case_qty is not None:
                lot_size = case_qty
            if lot_size is None:
                lot_size = 1

            conn.execute(sql, {
                "supplier_name": name,
                "jan": jan,
                "display_name": _to_str(cell("display_name")),
                "unit_price": _to_int(cell("unit_price")),
                "lot_size": lot_size,
                "case_qty": case_qty,
                "min_order_amount": _to_int(cell("min_order_amount")),
                "order_condition": _to_str(cell("order_condition")),
                "lead_time_text": _to_str(cell("lead_time_text")),
                "zone": zone,
                "zone_rank": zone_rank,
                "nst_supplier_code": nst_code,
                "source_sheet": name,
                "imported_at": now,
            })
            n_sheet += 1

        rows_ins += n_sheet
        sheets_done += 1

    conn.commit()
    wb.close()
    return {
        "sheets_processed": sheets_done,
        "rows_inserted": rows_ins,
        "skipped_no_jan": skipped_no_jan,
        "sheets_skipped": sheets_skipped,
        "warnings": warnings,
    }
