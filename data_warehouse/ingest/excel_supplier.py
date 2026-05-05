"""供应商管理 Excel ingestor — 仕入先管理リスト 导入。

处理：supplier_cost, supply_cycle, supplier_jan_list 表。
"""
from __future__ import annotations

import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def _to_float(value) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def assign_bucket(days) -> str:
    """根据 lead_time_days 分配 bucket。

    ≤7: 'short', 8-30: 'normal', 31-60: 'long', 其他/null: 'normal'
    """
    if days is None or not isinstance(days, (int, float)):
        return 'normal'
    days = int(days) if isinstance(days, float) else days
    if days <= 7:
        return 'short'
    if days <= 30:
        return 'normal'
    if days <= 60:
        return 'long'
    return 'normal'


class SupplierCostIngestor:
    """处理「仕入先-原価_AB類」和「仕入先_原価_C類」sheet → supplier_cost 表。"""

    def run(self, workbook_path: str, conn: sqlite3.Connection) -> dict:
        wb = load_workbook(workbook_path, data_only=True)
        inserted, errors = 0, 0
        now = datetime.now(timezone.utc).isoformat()

        for sheet_name, cost_class in [('仕入先-原価_AB類', 'AB'), ('仕入先_原価_C類', 'C')]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            for row_idx in range(5, ws.max_row + 1):
                jan = ws.cell(row_idx, 2).value
                if not jan:
                    continue
                jan = str(jan).strip()

                # Process main (col 11-12) and sub (col 14, 16) suppliers
                for supplier_col, cost_col in [(11, 12), (14, 16)]:
                    supplier = ws.cell(row_idx, supplier_col).value
                    if not supplier:
                        continue
                    supplier = str(supplier).strip()
                    cost = _to_float(ws.cell(row_idx, cost_col).value)
                    try:
                        conn.execute(
                            """INSERT OR REPLACE INTO supplier_cost
                               (jan, supplier_name, cost_class, unit_cost, currency, ingested_at)
                               VALUES (?, ?, ?, ?, ?, ?)""",
                            (jan, supplier, cost_class, cost, 'PHP', now)
                        )
                        inserted += 1
                    except Exception:
                        errors += 1

        conn.commit()
        return {'inserted': inserted, 'updated': 0, 'errors': errors}


class SupplyCycleIngestor:
    """处理「AB 商品进货周期」sheet → supply_cycle 表（含 bucket 自动分类）。"""

    def run(self, workbook_path: str, conn: sqlite3.Connection) -> dict:
        wb = load_workbook(workbook_path, data_only=True)
        if 'AB商品进货周期' not in wb.sheetnames:
            return {'inserted': 0, 'updated': 0, 'errors': 0}

        ws = wb['AB商品进货周期']
        inserted, errors = 0, 0
        now = datetime.now(timezone.utc).isoformat()

        for row_idx in range(2, ws.max_row + 1):
            jan = ws.cell(row_idx, 1).value
            if not jan:
                continue
            jan = str(jan).strip()

            lead_time_days = self._parse_lead_time(ws.cell(row_idx, 5).value)
            bucket = assign_bucket(lead_time_days)

            try:
                conn.execute(
                    """INSERT OR REPLACE INTO supply_cycle
                       (jan, lead_time_days, bucket, ingested_at)
                       VALUES (?, ?, ?, ?)""",
                    (jan, lead_time_days, bucket, now)
                )
                inserted += 1
            except Exception:
                errors += 1

        conn.commit()
        return {'inserted': inserted, 'updated': 0, 'errors': errors}

    @staticmethod
    def _parse_lead_time(text) -> int | None:
        """解析 lead_time 文本为天数。"""
        if not text:
            return None
        text = str(text).strip().lower()
        match = re.search(r'(\d+)', text)
        if not match:
            return None
        days = int(match.group(1))
        if '週' in text or 'week' in text:
            return days * 7
        if '个月' in text or 'ヶ月' in text or 'month' in text:
            return days * 30
        return days  # Default: interpret as days


class SupplierJanListIngestor:
    """处理供应商 sheet → supplier_jan_list 表（缺失 sheet 不报错）。"""

    def run(self, workbook_path: str, conn: sqlite3.Connection) -> dict:
        wb = load_workbook(workbook_path, data_only=True)
        inserted, errors = 0, 0
        now = datetime.now(timezone.utc).isoformat()

        for supplier_name in ['NEW WIND', '中央物産', '菅野', 'Maple']:
            if supplier_name not in wb.sheetnames:
                continue
            ws = wb[supplier_name]

            # Find header row (usually row 1)
            header_row = 1
            for row_idx in range(1, min(5, ws.max_row + 1)):
                cell = ws.cell(row_idx, 1).value
                if cell and str(cell).strip().upper() in ['JAN', 'SKU', '商品コード']:
                    header_row = row_idx
                    break

            for row_idx in range(header_row + 1, ws.max_row + 1):
                jan = ws.cell(row_idx, 1).value
                if not jan:
                    continue
                jan = str(jan).strip()
                try:
                    conn.execute(
                        """INSERT OR REPLACE INTO supplier_jan_list
                           (jan, supplier_name, status, ingested_at)
                           VALUES (?, ?, ?, ?)""",
                        (jan, supplier_name, None, now)
                    )
                    inserted += 1
                except Exception:
                    errors += 1

        conn.commit()
        return {'inserted': inserted, 'updated': 0, 'errors': errors}


def main(xlsx_path: str, db_path: str = "data_warehouse/warehouse.db"):
    """主入口：依次执行三个 ingestor。"""
    from data_warehouse.db.migrations import run as init_db

    xlsx_path = Path(xlsx_path).expanduser()
    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        return 1

    conn = init_db(db_path)
    print(f"Ingesting from {xlsx_path.name}...")

    for ingestor_cls, name in [
        (SupplierCostIngestor, 'supplier_cost'),
        (SupplyCycleIngestor, 'supply_cycle'),
        (SupplierJanListIngestor, 'supplier_jan_list'),
    ]:
        result = ingestor_cls().run(str(xlsx_path), conn)
        print(f"  {name}: {result['inserted']} inserted, {result['errors']} errors")

    conn.close()
    print("Done.")
    return 0


if __name__ == '__main__':
    import sys
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else ''))
