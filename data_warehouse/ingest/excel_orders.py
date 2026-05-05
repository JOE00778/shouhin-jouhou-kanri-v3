"""Shopee 订单数据导入（订单导出-*.xlsx）。

从 Sheet0 导入：
- 支付币种 / 单价 / 发货数量 / 本地SKU / 支付金额 / 平台 / 订单号 / 店铺
- 本地SKU 即 JAN（外键关联 item_master.jan）
"""
from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

import openpyxl


def _to_float(value) -> float | None:
    """尝试转为 float，失败返回 None。"""
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def ingest_orders(ws, conn: sqlite3.Connection) -> int:
    """Sheet0 → shopee_orders。

    header：支付币种 / 单价 / 发货数量 / 本地SKU / 支付金额 / 平台 / 订单号 / 店铺
    """
    # 读取 header
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_map = {col: idx for idx, col in enumerate(header_row)}

    # 找到各列的索引
    currency_idx = header_map.get("支付币种")
    unit_price_idx = header_map.get("单价")
    qty_idx = header_map.get("发货数量")
    jan_idx = header_map.get("本地SKU")
    payment_amount_idx = header_map.get("支付金额")
    platform_idx = header_map.get("平台")
    order_no_idx = header_map.get("订单号")
    shop_name_idx = header_map.get("店铺")

    inserted = 0
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    for row in rows:
        if not row or all(v is None for v in row):
            continue

        order_no = str(row[order_no_idx] or "").strip() if order_no_idx is not None and row and len(row) > order_no_idx and row[order_no_idx] is not None else None
        jan = str(row[jan_idx] or "").strip() if jan_idx is not None and row and len(row) > jan_idx and row[jan_idx] is not None else None

        if not order_no or not jan:
            continue

        try:
            currency = str(row[currency_idx] or "").strip() if currency_idx is not None and row and len(row) > currency_idx and row[currency_idx] is not None else None
            unit_price = _to_float(row[unit_price_idx] if unit_price_idx is not None and row and len(row) > unit_price_idx else None)
            qty = _to_float(row[qty_idx] if qty_idx is not None and row and len(row) > qty_idx else None)
            payment_amount = _to_float(row[payment_amount_idx] if payment_amount_idx is not None and row and len(row) > payment_amount_idx else None)
            platform = str(row[platform_idx] or "").strip() if platform_idx is not None and row and len(row) > platform_idx and row[platform_idx] is not None else None
            shop_name = str(row[shop_name_idx] or "").strip() if shop_name_idx is not None and row and len(row) > shop_name_idx and row[shop_name_idx] is not None else None

            conn.execute(
                """
                INSERT OR REPLACE INTO shopee_orders
                (order_no, sku_or_jan, qty, unit_price, payment_amount, currency, platform, shop_name)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (order_no, jan, qty, unit_price, payment_amount, currency, platform, shop_name),
            )
            inserted += 1
        except sqlite3.Error as e:
            print(f"  Warning: 订单记录写入失败 {order_no}/{jan}：{e}", file=sys.stderr)

    return inserted


def main(excel_path: str, db_path: str = "data_warehouse/warehouse.db"):
    """主入口：读 Excel Sheet0，执行 ingest，返回汇总。"""
    from data_warehouse.db.migrations import run as init_db

    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"错误：文件不存在 {excel_path}", file=sys.stderr)
        sys.exit(1)

    conn = init_db(db_path)

    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)

        print(f"[excel_orders] 开始导入 {excel_path.name}")
        print(f"  Sheets: {wb.sheetnames}")

        if "Sheet0" not in wb.sheetnames:
            print(f"  错误：未找到 Sheet0", file=sys.stderr)
            sys.exit(1)

        print(f"  处理 Sheet0...")
        count = ingest_orders(wb["Sheet0"], conn)
        print(f"    ✓ shopee_orders: {count} 条")

        conn.commit()
        print(f"[excel_orders] 完成，共入库 {count} 条记录")

    finally:
        conn.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"用法：{sys.argv[0]} <excel_path> [db_path]", file=sys.stderr)
        sys.exit(1)

    excel_path = sys.argv[1]
    db_path = sys.argv[2] if len(sys.argv) > 2 else "data_warehouse/warehouse.db"
    main(excel_path, db_path)
