"""6 个 NetSuite 导出 .xls 的 ingestor（函数式，每个文件一个函数）。

设计：
- 每个函数接受 (path, conn)，返回 {run_id, total, inserted, errors}
- 用 shared.xml_xls 解析 SpreadsheetML
- 共用 _start_run / _finalize_run / _record_error 落审计

涉及的表：
- inventory_snapshot — 库存数据（含 std_cost + avg_cost，喂 #1 成本同步）
- sales_line — 销售明细（4 类销售导出共用，source 字段区分）
- inventory_turnover — 库存周转率
"""
from __future__ import annotations

import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from shared.filters import ALLOWED_INVENTORY_LOCATIONS
from shared.xml_xls import iter_rows, parse_to_dicts


# ============================================================
# 通用工具
# ============================================================
def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def _to_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ============================================================
# Audit helpers（共用）
# ============================================================
def _start_run(conn: sqlite3.Connection, ingestor: str, source_file: str) -> int:
    cursor = conn.execute(
        """
        INSERT INTO _ingest_runs (ingestor, source_file, total_rows, inserted, updated, errors, run_at)
        VALUES (?, ?, 0, 0, 0, 0, ?)
        """,
        (ingestor, source_file, _now_iso()),
    )
    rid = cursor.lastrowid
    if rid is None:
        raise RuntimeError("无法获取 run_id")
    return rid


def _finalize_run(
    conn: sqlite3.Connection, run_id: int, *, total: int, inserted: int, errors: int
) -> None:
    conn.execute(
        "UPDATE _ingest_runs SET total_rows=?, inserted=?, errors=? WHERE run_id=?",
        (total, inserted, errors, run_id),
    )
    conn.commit()


def _record_error(
    conn: sqlite3.Connection, run_id: int, row_number: int, message: str, raw_row: dict
) -> None:
    conn.execute(
        "INSERT INTO _ingest_errors (run_id, row_number, error_message, raw_row) VALUES (?, ?, ?, ?)",
        (run_id, row_number, message, json.dumps(raw_row, ensure_ascii=False)),
    )


# Phase 4 · 通用 v2 helper
def _is_valid_jan(jan) -> bool:
    """JAN 必须是 8-13 位数字字符串。"""
    if not jan:
        return False
    s = str(jan).strip()
    return s.isdigit() and 8 <= len(s) <= 13


def _ensure_shop(conn, shop_id: str, market: str = "JP", platform: str = "unknown") -> None:
    """ingest 销售时自动 upsert shop 主档（避免 shop 表缺该 shop_id）。"""
    if not shop_id:
        return
    try:
        conn.execute(
            "INSERT OR REPLACE INTO shop "
            "(shop_id, market_id, platform, display_name, currency, owner, active, created_at) "
            "VALUES (?, ?, ?, ?, NULL, NULL, 1, ?) ",
            (shop_id, market, platform, shop_id, _now_iso()),
        )
    except Exception:
        pass


def _infer_shop_meta(shop_id: str) -> tuple[str, str]:
    """从 shop_id 字符串推断 (market, platform)。规则尽量保守。"""
    if not shop_id:
        return "JP", "unknown"
    s = shop_id.lower()
    # 平台
    if any(k in s for k in ("shopee", "smkj", "mtkshop")):
        platform = "shopee"
    elif "lazada" in s or "lzd" in s:
        platform = "lazada"
    elif "tokopedia" in s or "tpd" in s:
        platform = "tokopedia"
    elif "amazon" in s or "amzn" in s:
        platform = "amazon"
    elif "coupang" in s or "クーパン" in shop_id:
        platform = "coupang"
    elif "rakuten" in s or "楽天" in shop_id:
        platform = "rakuten"
    else:
        platform = "unknown"
    # 市场
    for code in ("tw", "sg", "my", "ph", "th", "vn", "id", "jp", "us", "kr", "cn", "br"):
        if (f" {code} " in f" {s} " or f"_{code}" in s or f"-{code}" in s
                or s.endswith(code) or s.startswith(code + "_")):
            return code.upper(), platform
    return "JP", platform


# ============================================================
# Period 解析（从 NetSuite 报表第 3 行 "2026年04月01日 - 2026年04月30日"）
# ============================================================
_PERIOD_PATTERN = re.compile(r"(\d{4})年(\d{1,2})月(\d{1,2})日\s*-\s*(\d{4})年(\d{1,2})月(\d{1,2})日")


def _extract_period(path: Path) -> tuple[str, str]:
    """从 NetSuite 报表 preamble 第 3 行提取期间。失败返回 ('', '')."""
    rows = list(iter_rows(path))
    for row in rows[:8]:
        for cell in row:
            if cell:
                m = _PERIOD_PATTERN.search(str(cell))
                if m:
                    y1, m1, d1, y2, m2, d2 = m.groups()
                    return (
                        f"{y1}-{int(m1):02d}-{int(d1):02d}",
                        f"{y2}-{int(m2):02d}-{int(d2):02d}",
                    )
    return ("", "")


# ============================================================
# Ingestor 1：inventory_snapshot（FB全倉庫通常在庫数残数検索結果）
# ============================================================
def ingest_inventory_snapshot(
    path: Path, conn: sqlite3.Connection, *, source_name: str | None = None
) -> dict:
    """直写 item_inventory_snapshot_v2（Phase 4.2）· jan 强制 8-13 位数字。

    数据流：xls → item_inventory_snapshot_v2（覆盖快照）+ 同步 item_v2.汇总字段
    """
    path = Path(path)
    source_name = source_name or path.name
    snapshot_at = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()

    run_id = _start_run(conn, "item_inventory_snapshot_v2", source_name)
    inserted = errors = skipped_no_jan = skipped_other_loc = 0

    # 整表覆盖（最新快照）
    conn.execute("DELETE FROM item_inventory_snapshot_v2")

    rows = parse_to_dicts(path, header_row=0)
    sql = """
        INSERT OR REPLACE INTO item_inventory_snapshot_v2 (
            jan, item_code, internal_id, display_name,
            location, bin_number, snapshot_at,
            qty_on_hand, qty_committed, qty_backorder,
            std_cost, avg_cost, total_amount,
            handling_status, status, owner, department,
            imported_at
        ) VALUES (
            :jan, :item_code, :internal_id, :display_name,
            :location, :bin_number, :snapshot_at,
            :qty_on_hand, :qty_committed, :qty_backorder,
            :std_cost, :avg_cost, :total_amount,
            :handling_status, :status, :owner, :department,
            :imported_at
        )
    """
    now = _now_iso()

    for n, raw in enumerate(rows, start=1):
        try:
            jan = _to_str(raw.get("UPCコード"))
            internal_id = _to_str(raw.get("内部ID"))
            item_code = _to_str(raw.get("アイテム"))
            # 跳过空行 / 汇总行
            if not internal_id or not item_code:
                continue
            if internal_id in ("総合計", "合計"):
                continue
            # JAN 强制
            if not _is_valid_jan(jan):
                skipped_no_jan += 1
                continue
            # 仓库白名单
            location = _to_str(raw.get("場所"))
            if location not in ALLOWED_INVENTORY_LOCATIONS:
                skipped_other_loc += 1
                continue

            payload = {
                "jan": jan,
                "item_code": item_code,
                "internal_id": internal_id,
                "display_name": _to_str(raw.get("表示名")),
                "location": location,
                "bin_number": _to_str(raw.get("保管棚番号")),
                "snapshot_at": snapshot_at,
                "qty_on_hand": _to_float(raw.get("手持合計")),
                "qty_committed": _to_float(raw.get("確保済合計")),
                "qty_backorder": _to_float(raw.get("バック・オーダー合計")),
                "std_cost": _to_float(raw.get("アイテム定義原価")),
                "total_amount": _to_float(raw.get("合計金額")),
                "avg_cost": _to_float(raw.get("平均原価合計")),
                "handling_status": _to_str(raw.get("取扱区分")),
                "status": _to_str(raw.get("ステータス")),
                "owner": _to_str(raw.get("商品担当者")),
                "department": _to_str(raw.get("部門")),
                "imported_at": now,
            }
            conn.execute(sql, payload)
            inserted += 1
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            errors += 1
            try:
                _record_error(conn, run_id, n, str(e), raw)
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass

    # 同步 item_v2 库存汇总字段（按 jan 聚合）
    try:
        conn.execute(
            """
            UPDATE item_v2 SET
              on_hand_total = COALESCE((
                SELECT SUM(qty_on_hand) FROM item_inventory_snapshot_v2 i WHERE i.jan = item_v2.jan
              ), 0),
              qty_committed_total = COALESCE((
                SELECT SUM(qty_committed) FROM item_inventory_snapshot_v2 i WHERE i.jan = item_v2.jan
              ), 0),
              total_amount = COALESCE((
                SELECT SUM(total_amount) FROM item_inventory_snapshot_v2 i WHERE i.jan = item_v2.jan
              ), 0),
              updated_at = ?
            WHERE jan IN (SELECT DISTINCT jan FROM item_inventory_snapshot_v2)
            """,
            (now,),
        )
    except Exception:
        pass  # item_v2 还没建（首次 ingest）则跳过

    _finalize_run(conn, run_id, total=len(rows), inserted=inserted, errors=errors)
    return {
        "run_id": run_id, "total": len(rows),
        "inserted": inserted, "errors": errors,
        "period_start": None, "period_end": None,
    }


# ============================================================
# Ingestor 1b：inventory_snapshot_multi（在庫のスナップショット-980 · 多仓多级表头）
# ============================================================
# 子表头按仓库循环出现的 8 列字段
_MULTI_SUB_FIELDS = (
    "適正在庫水準", "手持", "注文済", "確保済",
    "注文待ち", "輸送中", "平均原価", "定義原価",
)
# 主表头前 6 列（物品级元数据）
_MULTI_META_COLS = ("内部ID", "UPCコード", "在庫アイテム: 表示名", "ランク", "取扱区分", "保管棚番号")


def _parse_multi_warehouse_header(rows: list[list]) -> tuple[list[str], dict[str, dict[str, int]]]:
    """从 row 6 (主表头) + row 7 (子表头) 解析仓库列表 + 列索引映射。

    返回:
      warehouses: 仓库名顺序列表（含「合計」, 末尾）
      col_index: { warehouse_name: { sub_field_name: col_idx } }
    """
    main = rows[6]
    sub = rows[7]
    warehouses: list[str] = []
    col_index: dict[str, dict[str, int]] = {}

    # 仓库块从 col 6 起（前 6 列是物品级 meta），每块占 8 列
    n_meta = len(_MULTI_META_COLS)  # 6
    block_size = len(_MULTI_SUB_FIELDS)  # 8

    col = n_meta
    while col < len(main):
        wh = main[col]
        if wh:
            warehouses.append(str(wh).strip())
            col_index[str(wh).strip()] = {
                fld: col + i for i, fld in enumerate(_MULTI_SUB_FIELDS) if (col + i) < len(sub)
            }
        col += block_size
    return warehouses, col_index


def ingest_inventory_snapshot_multi(
    path, conn: sqlite3.Connection, *, source_name: str | None = None
) -> dict:
    """新格式 在庫のスナップショット-980.xls · 多仓库多级表头.

    - 解析 row 6 (主表头, 仓库列表) + row 7 (子表头, 8 子字段)
    - 写入 item_inventory_snapshot_v2: 每个 (jan, location) 一行（不含合計行）
    - total_amount = 平均原価 × (手持 + 注文待ち + 輸送中)
    - 同步 item_v2.{on_hand_total, on_order_total, qty_committed_total, total_amount}
    """
    path = Path(path)
    source_name = source_name or path.name
    snapshot_at = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()

    run_id = _start_run(conn, "item_inventory_snapshot_v2.multi", source_name)
    inserted = errors = skipped_no_jan = skipped_other_loc = 0

    # 整表覆盖（最新快照）
    conn.execute("DELETE FROM item_inventory_snapshot_v2")

    rows = list(iter_rows(path))
    if len(rows) < 9:
        _finalize_run(conn, run_id, total=0, inserted=0, errors=1)
        return {
            "run_id": run_id, "total": 0, "inserted": 0, "errors": 1,
            "period_start": None, "period_end": None,
        }

    warehouses, col_index = _parse_multi_warehouse_header(rows)
    # 数据仓库列表（去掉「合計」）
    data_warehouses = [w for w in warehouses if w != "合計"]

    sql = """
        INSERT OR REPLACE INTO item_inventory_snapshot_v2 (
            jan, item_code, internal_id, display_name,
            location, bin_number, snapshot_at,
            qty_on_hand, qty_committed, qty_backorder,
            std_cost, avg_cost, total_amount,
            handling_status, status, owner, department,
            imported_at
        ) VALUES (
            :jan, :item_code, :internal_id, :display_name,
            :location, :bin_number, :snapshot_at,
            :qty_on_hand, :qty_committed, :qty_backorder,
            :std_cost, :avg_cost, :total_amount,
            :handling_status, :status, :owner, :department,
            :imported_at
        )
    """
    now = _now_iso()

    # item_v2 汇总：用「合計」列拿数量类（NetSuite 已聚合）, total_amount 用每仓 avg×qty 后再 SUM
    # 因「合計」列的 平均原価 是各仓 SUM(非真平均), 所以 total_amount 必须 per-warehouse 算后求和
    item_totals: dict[str, dict] = {}  # jan -> {on_hand, on_order, committed, total_amount}
    total_idx = col_index.get("合計", {})

    # 遍历数据行（从 row 8 起）
    total_data_rows = 0
    for n, raw in enumerate(rows[8:], start=9):
        try:
            internal_id = _to_str(raw[0]) if len(raw) > 0 else None
            jan = _to_str(raw[1]) if len(raw) > 1 else None
            display_name = _to_str(raw[2]) if len(raw) > 2 else None
            rank = _to_str(raw[3]) if len(raw) > 3 else None
            handling_status = _to_str(raw[4]) if len(raw) > 4 else None
            bin_number = _to_str(raw[5]) if len(raw) > 5 else None

            # 跳过空行 / 合計行
            if not internal_id:
                continue
            if internal_id in ("総合計", "合計", "合 計"):
                continue
            total_data_rows += 1
            # JAN 强制 8-13 位数字
            if not _is_valid_jan(jan):
                skipped_no_jan += 1
                continue

            # 用「合計」列计算 item_v2 汇总（每 jan 累加保险，但这文件 jan 唯一）
            def _cell(idx_map: dict[str, int], fld: str):
                idx = idx_map.get(fld)
                if idx is None or idx >= len(raw):
                    return None
                return _to_float(raw[idx])

            # 数量类用「合計」列（NetSuite 已聚合）
            tot_oh = _cell(total_idx, "手持") or 0.0
            tot_oo = _cell(total_idx, "注文済") or 0.0
            tot_cm = _cell(total_idx, "確保済") or 0.0
            # 金额类: per-warehouse avg × qty 后求和（合計 列的平均原価 是各仓 SUM 不可直用）
            jan_total_amt = 0.0
            for wh in data_warehouses:
                idx_map = col_index.get(wh, {})
                wh_avg = _cell(idx_map, "平均原価") or 0.0
                wh_oh = _cell(idx_map, "手持") or 0.0
                wh_wt = _cell(idx_map, "注文待ち") or 0.0
                wh_tr = _cell(idx_map, "輸送中") or 0.0
                jan_total_amt += wh_avg * (wh_oh + wh_wt + wh_tr)
            item_totals[jan] = {
                "on_hand_total": tot_oh,
                "on_order_total": tot_oo,
                "qty_committed_total": tot_cm,
                "total_amount": jan_total_amt,
            }

            # 每仓库一行（白名单内）
            for wh in data_warehouses:
                if wh not in ALLOWED_INVENTORY_LOCATIONS:
                    skipped_other_loc += 1
                    continue
                idx_map = col_index.get(wh, {})
                qty_on_hand = _cell(idx_map, "手持")
                qty_committed = _cell(idx_map, "確保済")
                qty_waiting = _cell(idx_map, "注文待ち")
                qty_transit = _cell(idx_map, "輸送中")
                avg_cost = _cell(idx_map, "平均原価")
                std_cost = _cell(idx_map, "定義原価")
                # backorder 在新文件不存在; 用「注文待ち」近似（语义最接近）
                qty_backorder = qty_waiting

                # total_amount = 平均原価 × (手持 + 注文待ち + 輸送中)
                qty_for_amt = (qty_on_hand or 0.0) + (qty_waiting or 0.0) + (qty_transit or 0.0)
                total_amount = (avg_cost or 0.0) * qty_for_amt if avg_cost else None

                payload = {
                    "jan": jan,
                    "item_code": None,  # 新文件无 アイテム 列
                    "internal_id": internal_id,
                    "display_name": display_name,
                    "location": wh,
                    "bin_number": bin_number or "",
                    "snapshot_at": snapshot_at,
                    "qty_on_hand": qty_on_hand,
                    "qty_committed": qty_committed,
                    "qty_backorder": qty_backorder,
                    "std_cost": std_cost,
                    "avg_cost": avg_cost,
                    "total_amount": total_amount,
                    "handling_status": handling_status,
                    "status": rank,  # ランク 放 status 暂存
                    "owner": None,
                    "department": None,
                    "imported_at": now,
                }
                conn.execute(sql, payload)
                inserted += 1
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            errors += 1
            try:
                _record_error(conn, run_id, n, str(e), {"row": list(raw[:6]) if raw else []})
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass

    # 同步 item_v2 汇总字段（用「合計」列直接算的值，per jan）
    if item_totals:
        try:
            for jan, agg in item_totals.items():
                conn.execute(
                    """
                    UPDATE item_v2 SET
                      on_hand_total = ?,
                      on_order_total = ?,
                      qty_committed_total = ?,
                      total_amount = ?,
                      updated_at = ?
                    WHERE jan = ?
                    """,
                    (
                        agg["on_hand_total"],
                        agg["on_order_total"],
                        agg["qty_committed_total"],
                        agg["total_amount"],
                        now,
                        jan,
                    ),
                )
        except Exception:
            pass  # item_v2 不存在则跳过

    _finalize_run(conn, run_id, total=total_data_rows, inserted=inserted, errors=errors)
    return {
        "run_id": run_id, "total": total_data_rows,
        "inserted": inserted, "errors": errors,
        "skipped_no_jan": skipped_no_jan,
        "skipped_other_loc": skipped_other_loc,
        "warehouses": data_warehouses,
        "period_start": None, "period_end": None,
    }


# ============================================================
# Ingestor 2-5：sales_line（4 个销售导出，source 不同）
# ============================================================
_STORE_PREFIXES = ("Shopee", "Lazada", "Tokopedia")
_STORE_KEYWORDS = ("COUPANG", "Coupang", "coupang")


def _is_store_group_header(item_code: str | None, display_name: str | None) -> bool:
    """判断一行是不是「店铺分组标题」（item_code 是店铺名，display_name 空）。

    放宽策略：只要 item_code 非空 + display_name 空 + item_code 不是纯数字（SKU 编码通常是 4901... 之类），
    就视为店铺标题。避免硬编码店铺前缀漏匹配（TikTok / Amazon / その他 等）。
    合計/小計行已在调用方单独过滤。
    """
    if not item_code or display_name:
        return False
    s = item_code.strip()
    # 纯数字（SKU 编码 / EAN / JAN）→ 不是店铺
    if s.isdigit():
        return False
    # 已知前缀 / 关键字（保留，命中即返回）
    if s.startswith(_STORE_PREFIXES):
        return True
    if any(k in s for k in _STORE_KEYWORDS):
        return True
    # 兜底：item_code 含字母 / 含空格 / 长度 > 4 → 多半是店铺名
    if any(c.isalpha() for c in s) or " " in s:
        return True
    return False


def _ingest_sales(
    path: Path,
    conn: sqlite3.Connection,
    *,
    source: str,
    granularity: str,               # 'monthly' / 'daily' / 'cumulative'
    has_store_column: bool,        # CSV 列里直接有 FB_店舗
    has_store_groups: bool,         # CSV 用「店铺标题行 + SKU 明细行」分组结构
    has_rank: bool,
    has_purchase_price: bool,
    has_upc: bool = True,
    source_name: str | None = None,
) -> dict:
    """通用销售导入 · 直写 shop_sales（Phase 4.2）。

    数据流：xls → shop_sales(granularity, period, shop_id, jan)
    JAN 强制 8-13 位数字；空行 / 汇总行 / 无 JAN 行跳过。
    """
    path = Path(path)
    source_name = source_name or path.name
    period_start, period_end = _extract_period(path)

    run_id = _start_run(conn, f"shop_sales.{source}", source_name)
    inserted = errors = skipped_no_jan = 0

    # 先删除同 (source, period, granularity) 的旧数据
    conn.execute(
        "DELETE FROM shop_sales WHERE source = ? AND period_start = ? AND period_end = ? AND granularity = ?",
        (source, period_start, period_end, granularity),
    )

    rows = parse_to_dicts(path, header_row=6)
    sql = """
        INSERT OR REPLACE INTO shop_sales (
            shop_id, jan, granularity, period_start, period_end,
            qty_sold, unit_price, revenue, revenue_jpy,
            cost, gross_profit, gross_margin, rank, source, imported_at
        ) VALUES (
            :shop_id, :jan, :granularity, :period_start, :period_end,
            :qty_sold, :unit_price, :revenue, :revenue_jpy,
            :cost, :gross_profit, :gross_margin, :rank, :source, :imported_at
        )
    """
    now = _now_iso()
    current_store: str | None = None
    seen_shops: set[str] = set()

    for n, raw in enumerate(rows, start=1):
        try:
            item_code = _to_str(raw.get("アイテム"))
            display_name = _to_str(raw.get("表示名"))

            # 店铺分组标题
            if _is_store_group_header(item_code, display_name):
                if has_store_groups:
                    current_store = item_code
                continue

            if not item_code:
                continue
            if any(k in item_code for k in ("合計", "合计", "総合計", "総計")):
                continue

            # JAN 强制（无 UPC 报表 fallback 用 item_code 如果是 13 位）
            upc = _to_str(raw.get("UPCコード"))
            if has_upc:
                jan = upc
            else:
                jan = upc if _is_valid_jan(upc) else item_code
            if not _is_valid_jan(jan):
                skipped_no_jan += 1
                continue

            # 决定 shop_id
            if has_store_column:
                store = _to_str(raw.get("FB_店舗"))
            elif has_store_groups:
                store = current_store
            else:
                store = None
            shop_id = store or f"netsuite_{source}"

            # 自动建 shop 主档（一个 source 内只建一次）
            if shop_id not in seen_shops:
                market, platform = _infer_shop_meta(shop_id)
                _ensure_shop(conn, shop_id, market=market, platform=platform)
                seen_shops.add(shop_id)

            revenue = _to_float(raw.get("総収益"))
            payload = {
                "shop_id": shop_id,
                "jan": jan,
                "granularity": granularity,
                "period_start": period_start or "",
                "period_end": period_end or "",
                "qty_sold": _to_float(raw.get("販売数量")),
                "unit_price": _to_float(raw.get("購入価格(単価)")) if has_purchase_price else None,
                "revenue": revenue,
                "revenue_jpy": revenue,  # 当地币 = JPY 假设（NetSuite ASEAN 报表已折算）
                "cost": _to_float(raw.get("定義原価")),
                "gross_profit": _to_float(raw.get("粗利")),
                "gross_margin": _to_float(raw.get("粗利率")),
                "rank": _to_str(raw.get("商品ランク")) if has_rank else None,
                "source": source,
                "imported_at": now,
            }
            conn.execute(sql, payload)
            inserted += 1
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            errors += 1
            try:
                _record_error(conn, run_id, n, str(e), raw)
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass

    _finalize_run(conn, run_id, total=len(rows), inserted=inserted, errors=errors)
    if len(rows) > 0 and inserted == 0:
        first_keys = list(rows[0].keys()) if rows else []
        raise RuntimeError(
            f"读取到 {len(rows)} 行但 0 行入库 (skipped_no_jan={skipped_no_jan})。"
            f"列名：{first_keys}。期待 アイテム / UPCコード / 販売数量 / 総収益 / 粗利"
        )
    return {
        "run_id": run_id, "total": len(rows),
        "inserted": inserted, "errors": errors,
        "period_start": period_start, "period_end": period_end,
    }


def ingest_sales_asean_monthly(path, conn, **kw):
    """ASEAN 店舗別売上 集計専用 → shop_sales (granularity='monthly')"""
    return _ingest_sales(
        path, conn, source="asean_monthly", granularity="monthly",
        has_store_column=True, has_store_groups=False,
        has_rank=False, has_purchase_price=False, **kw,
    )


def ingest_sales_asean_daily(path, conn, **kw):
    """ASEAN 店舗別売上(前日) → shop_sales (granularity='daily')"""
    return _ingest_sales(
        path, conn, source="asean_daily", granularity="daily",
        has_store_column=False, has_store_groups=True,
        has_rank=False, has_purchase_price=False, has_upc=False, **kw,
    )


def ingest_sales_export_item(path, conn, **kw):
    """輸出 アイテム別売上 → shop_sales (granularity='monthly', shop_id=netsuite_export_item)"""
    return _ingest_sales(
        path, conn, source="export_item", granularity="monthly",
        has_store_column=False, has_store_groups=False,
        has_rank=True, has_purchase_price=True, **kw,
    )


def ingest_sales_export_store(path, conn, **kw):
    """輸出 店舗別売上 → shop_sales (granularity='monthly')"""
    return _ingest_sales(
        path, conn, source="export_store", granularity="monthly",
        has_store_column=True, has_store_groups=False,
        has_rank=True, has_purchase_price=False, **kw,
    )


# ============================================================
# Ingestor 6：inventory_turnover（在庫回転率）
# ============================================================
def ingest_inventory_turnover(
    path, conn, *, source_name: str | None = None
) -> dict:
    """在庫回転率：(item_code, period) UPSERT。"""
    path = Path(path)
    source_name = source_name or path.name
    period_start, period_end = _extract_period(path)

    run_id = _start_run(conn, "inventory_turnover", source_name)
    inserted = 0
    errors = 0

    # 同期间重新上传 → 先删旧再插,避免累计
    conn.execute(
        "DELETE FROM inventory_turnover WHERE period_start = ? AND period_end = ?",
        (period_start, period_end),
    )

    rows = parse_to_dicts(path, header_row=6)
    sql = """
        INSERT OR REPLACE INTO inventory_turnover (
            item_code, description, cost, avg_value, turnover_rate, avg_days_on_hand,
            period_start, period_end, source_file, imported_at
        ) VALUES (
            :item_code, :description, :cost, :avg_value, :turnover_rate, :avg_days_on_hand,
            :period_start, :period_end, :source_file, :imported_at
        )
    """
    now = _now_iso()
    for n, raw in enumerate(rows, start=1):
        try:
            item_code = _to_str(raw.get("アイテム"))
            if not item_code or item_code in ("在庫アイテム", "合計", "総合計"):
                continue  # 跳过分组标题行
            payload = {
                "item_code": item_code,
                "description": _to_str(raw.get("説明")),
                "cost": _to_float(raw.get("原価")),
                "avg_value": _to_float(raw.get("平均値")),
                "turnover_rate": _to_float(raw.get("回転率")),
                "avg_days_on_hand": _to_float(raw.get("平均手持日数")),
                "period_start": period_start,
                "period_end": period_end,
                "source_file": source_name,
                "imported_at": now,
            }
            conn.execute(sql, payload)
            inserted += 1
        except Exception as e:
            errors += 1
            _record_error(conn, run_id, n, str(e), raw)

    _finalize_run(conn, run_id, total=len(rows), inserted=inserted, errors=errors)
    return {
        "run_id": run_id,
        "total": len(rows),
        "inserted": inserted,
        "errors": errors,
        "period_start": period_start,
        "period_end": period_end,
    }


# ============================================================
# Ingestor 7：shopee_orders_raw（订单导出.xlsx Sheet0）
# ============================================================
def ingest_shopee_orders_raw(
    path, conn, *, source_name: str | None = None
) -> dict:
    """订单导出 .xlsx Sheet0 (8 列): 支付币种/单价/发货数量/本地SKU/支付金额/平台/订单号/店铺."""
    import openpyxl
    path = Path(path)
    source_name = source_name or path.name
    run_id = _start_run(conn, "shopee_orders_raw", source_name)

    # 整表 truncate 再插(只更新不累积)
    conn.execute("DELETE FROM shopee_orders_raw")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = list(ws.iter_rows(values_only=True))
    if len(rows_iter) < 2:
        _finalize_run(conn, run_id, total=0, inserted=0, errors=0)
        return {"run_id": run_id, "total": 0, "inserted": 0, "errors": 0,
                "period_start": None, "period_end": None}

    header = [str(h).strip() if h is not None else "" for h in rows_iter[0]]
    # 字段映射 (中文表头 → DB 字段)
    name_map = {
        "支付币种": "currency", "单价": "unit_price", "发货数量": "ship_qty",
        "本地SKU": "local_sku", "支付金额": "payment_amount",
        "平台": "platform", "订单号": "order_no", "店铺": "shop_name",
    }
    col_to_field = {i: name_map[h] for i, h in enumerate(header) if h in name_map}

    sql = """
        INSERT OR REPLACE INTO shopee_orders_raw (
            currency, unit_price, ship_qty, local_sku, payment_amount,
            platform, order_no, shop_name, source_file, imported_at
        ) VALUES (
            :currency, :unit_price, :ship_qty, :local_sku, :payment_amount,
            :platform, :order_no, :shop_name, :source_file, :imported_at
        )
    """
    now = _now_iso()
    inserted = errors = 0
    total = len(rows_iter) - 1
    for n, row in enumerate(rows_iter[1:], start=1):
        try:
            payload = {
                "currency": None, "unit_price": None, "ship_qty": None,
                "local_sku": None, "payment_amount": None,
                "platform": None, "order_no": None, "shop_name": None,
                "source_file": source_name, "imported_at": now,
            }
            for i, field in col_to_field.items():
                if i < len(row):
                    v = row[i]
                    if field == "payment_amount":
                        payload[field] = _to_float(v)
                    else:
                        payload[field] = _to_str(v)
            if not payload["order_no"]:
                continue  # 跳过空行
            conn.execute(sql, payload)
            inserted += 1
        except Exception as e:
            errors += 1
            _record_error(conn, run_id, n, str(e), {"row": str(row)[:200]})

    conn.commit()
    _finalize_run(conn, run_id, total=total, inserted=inserted, errors=errors)
    return {"run_id": run_id, "total": total, "inserted": inserted,
            "errors": errors, "period_start": None, "period_end": None}


# ============================================================
# Ingestor 8：shopee_income_lines（ph.mtkshop.*.income.已拨款.xlsx Income sheet）
# ============================================================
def ingest_shopee_income(
    path, conn, *, source_name: str | None = None
) -> dict:
    """ph.*.income.已拨款.xlsx Income sheet (R6 表头, 46 列)."""
    import openpyxl
    path = Path(path)
    source_name = source_name or path.name
    run_id = _start_run(conn, "shopee_income", source_name)

    wb = openpyxl.load_workbook(str(path), data_only=True)
    if "Income" not in wb.sheetnames:
        _finalize_run(conn, run_id, total=0, inserted=0, errors=0)
        return {"run_id": run_id, "total": 0, "inserted": 0, "errors": 0,
                "period_start": None, "period_end": None}
    ws = wb["Income"]

    # 同 (seller_account, payout_date) 重新上传 → 先删旧再插,避免累计
    # 注: 提前读 R2 D 列拿 payout_date,再 DELETE
    rows_pre = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    seller_pre = str(rows_pre[0][0]).strip() if rows_pre and rows_pre[0][0] else None
    payout_pre = (
        str(rows_pre[0][3]).strip().split(" ")[0].split("T")[0]
        if rows_pre and len(rows_pre[0]) > 3 and rows_pre[0][3] else None
    )
    if seller_pre and payout_pre:
        conn.execute(
            "DELETE FROM shopee_income_lines WHERE seller_account = ? AND payout_date = ?",
            (seller_pre, payout_pre),
        )

    # R1: 卖家帐号/付款ID/收款渠道/拨款时间 (header)
    # R2: 值 (mtkshop.ph / 2026-04-01 等)
    # R6: 主表头
    # R7+: detail
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 7:
        _finalize_run(conn, run_id, total=0, inserted=0, errors=0)
        return {"run_id": run_id, "total": 0, "inserted": 0, "errors": 0,
                "period_start": None, "period_end": None}

    # 顶部 meta (R2)
    seller_account = _to_str(rows[1][0]) if len(rows[1]) > 0 else None
    payout_date_top = _to_str(rows[1][3]) if len(rows[1]) > 3 else None
    if payout_date_top:
        # 'YYYY-MM-DD HH:MM:SS' 或 'YYYY-MM-DD'
        payout_date_top = payout_date_top.split(" ")[0].split("T")[0]

    # 表头映射 (中文 → DB 字段)
    header = [str(h).strip() if h is not None else "" for h in rows[5]]  # R6 (0-indexed=5)
    NAME_MAP = {
        "编号": "seq", "订单编号": "order_no", "退款ID": "refund_id",
        "买家帐号": "buyer_account", "订单成立时间": "order_created_at",
        "买家付款方式": "payment_method", "Hot Listing": "hot_listing",
        "买家付款方式详情_1": "payment_method_detail",
        "分期付款计划 （如适用）": "installment_plan",
        "installment rate": "installment_rate",
        "拨款完成日期": "payout_completed_at",
        "商品原价": "gross_price", "商品折扣": "product_discount",
        "退款金額": "refund_amount", "Shopee回扣金额": "shopee_rebate",
        "卖家赞助的优惠券": "seller_voucher",
        "卖家赞助的合资优惠券": "seller_voucher_jv",
        "卖家赞助的 Shopee 币回扣": "seller_shopee_coin",
        "卖家赞助的合资 Shopee 币回扣": "seller_shopee_coin_jv",
        "买家支付运费": "buyer_shipping",
        "Shopee运费补贴": "shopee_shipping_subsidy",
        "卖家支付运费": "seller_shipping",
        "退货运费": "return_shipping",
        "退货给卖家的运费": "return_to_seller_ship",
        "通过运费险计划节省下的运费总额": "shipping_insurance_save",
        "联盟营销方案佣金": "affiliate_commission",
        "佣金": "commission",
        "物流+：海外免退服务-派送失败场景服务费": "fbs_overseas_fail",
        "物流+：海外免退服务-退货退款场景服务费": "fbs_overseas_return",
        "服务费": "service_fee",
        "运费险计划活动服务费": "shipping_insurance_fee",
        "交易手续费": "transaction_fee",
        "FBS Fee": "fbs_fee",
        "拨款金额 (₱)": "payout_amount",
        "优惠码": "promo_code",
        "损失赔偿": "loss_compensation",
        "每个订单的实际总重量": "actual_weight",
        "卖家提供的运费促销": "seller_shipping_promo",
        "物流承运商": "logistics_carrier",
        "物流名称": "logistics_name",
        "退款给买家的现金金额": "refund_cash",
        "退货/退款商品的按比例Shopee币抵消": "prorated_shopee_coin",
        "退货商品的按比例Shopee优惠券抵消": "prorated_shopee_voucher",
        "Pro-rated Bank Payment Channel Promotion  for return refund Items": "prorated_bank_promo",
        "Pro-rated Shopee Payment Channel Promotion  for return refund Items": "prorated_payment_promo",
    }
    col_to_field = {i: NAME_MAP[h] for i, h in enumerate(header) if h in NAME_MAP}
    NUMERIC_FIELDS = {
        "gross_price", "product_discount", "refund_amount", "shopee_rebate",
        "seller_voucher", "seller_voucher_jv", "seller_shopee_coin",
        "seller_shopee_coin_jv", "buyer_shipping", "shopee_shipping_subsidy",
        "seller_shipping", "return_shipping", "return_to_seller_ship",
        "shipping_insurance_save", "affiliate_commission", "commission",
        "fbs_overseas_fail", "fbs_overseas_return", "service_fee",
        "shipping_insurance_fee", "transaction_fee", "fbs_fee",
        "payout_amount", "loss_compensation", "actual_weight",
        "seller_shipping_promo", "refund_cash",
        "prorated_shopee_coin", "prorated_shopee_voucher",
        "prorated_bank_promo", "prorated_payment_promo",
    }
    INT_FIELDS = {"seq"}

    # 所有 DB 字段（按 schema 顺序）
    DB_FIELDS = [
        "seq", "order_no", "refund_id", "buyer_account", "order_created_at",
        "payment_method", "hot_listing", "payment_method_detail",
        "installment_plan", "installment_rate", "payout_completed_at",
        "gross_price", "product_discount", "refund_amount", "shopee_rebate",
        "seller_voucher", "seller_voucher_jv", "seller_shopee_coin",
        "seller_shopee_coin_jv", "buyer_shipping", "shopee_shipping_subsidy",
        "seller_shipping", "return_shipping", "return_to_seller_ship",
        "shipping_insurance_save", "affiliate_commission", "commission",
        "fbs_overseas_fail", "fbs_overseas_return", "service_fee",
        "shipping_insurance_fee", "transaction_fee", "fbs_fee",
        "payout_amount", "promo_code", "loss_compensation", "actual_weight",
        "seller_shipping_promo", "logistics_carrier", "logistics_name",
        "refund_cash", "prorated_shopee_coin", "prorated_shopee_voucher",
        "prorated_bank_promo", "prorated_payment_promo",
        "seller_account", "payout_date", "source_file", "imported_at",
    ]
    placeholders = ",".join(f":{f}" for f in DB_FIELDS)
    cols = ",".join(DB_FIELDS)
    sql = f"INSERT OR REPLACE INTO shopee_income_lines ({cols}) VALUES ({placeholders})"

    now = _now_iso()
    inserted = errors = 0
    detail_rows = rows[6:]  # R7+ (0-indexed=6)
    total = 0
    for n, row in enumerate(detail_rows, start=1):
        try:
            if not row or all(v is None for v in row):
                continue
            payload = {f: None for f in DB_FIELDS}
            payload["seller_account"] = seller_account
            payload["payout_date"] = payout_date_top
            payload["source_file"] = source_name
            payload["imported_at"] = now

            for i, field in col_to_field.items():
                if i < len(row):
                    v = row[i]
                    if field in NUMERIC_FIELDS:
                        payload[field] = _to_float(v)
                    elif field in INT_FIELDS:
                        try:
                            payload[field] = int(v) if v is not None else None
                        except (ValueError, TypeError):
                            payload[field] = None
                    else:
                        payload[field] = _to_str(v)

            if not payload.get("order_no"):
                continue  # 跳过空行 / 小计行
            total += 1
            conn.execute(sql, payload)
            inserted += 1
        except Exception as e:
            errors += 1
            _record_error(conn, run_id, n, str(e), {"row": str(row)[:200]})

    conn.commit()
    _finalize_run(conn, run_id, total=total, inserted=inserted, errors=errors)
    return {"run_id": run_id, "total": total, "inserted": inserted,
            "errors": errors, "period_start": payout_date_top, "period_end": payout_date_top}


def ingest_item_summary(path, conn, *, source_name: str | None = None) -> dict:
    """アイテム.xls 8 列 → 直写 item_v2（Phase 4.2）。

    8 列对应 アイテム.xls：A=item_code, B=upc(jan), C=display_name,
        D=handling_status, E=std_cost, F=available, G=available_on_hand, H=avg_cost
    JAN 强制 8-13 位数字；既有 jan UPSERT 更新 std_cost / avg_cost / handling_status。
    """
    from data_warehouse.ingest.xml_netsuite import ItemSummaryIngestor
    path = Path(path)
    source_name = source_name or path.name
    run_id = _start_run(conn, "item_v2.item_summary", source_name)
    inserted = errors = skipped_no_jan = 0

    # 用 ItemSummaryIngestor 解析 XML 结构
    parser = ItemSummaryIngestor()
    rows = parser.parse_rows(str(path))

    sql = """
        INSERT INTO item_v2 (
            jan, item_code, upc, display_name, handling_status,
            std_cost, avg_cost, source_priority, imported_at, updated_at
        ) VALUES (
            :jan, :item_code, :upc, :display_name, :handling_status,
            :std_cost, :avg_cost, 'nst_item_summary', :now, :now
        )
        ON CONFLICT (jan) DO UPDATE SET
          item_code = EXCLUDED.item_code,
          display_name = EXCLUDED.display_name,
          handling_status = EXCLUDED.handling_status,
          std_cost = EXCLUDED.std_cost,
          avg_cost = EXCLUDED.avg_cost,
          updated_at = EXCLUDED.updated_at
    """
    now = _now_iso()
    for n, raw in enumerate(rows, start=1):
        try:
            payload = parser.parse_row(raw)
            if payload is None:
                continue
            jan = (payload.get("upc") or "").strip()
            if not _is_valid_jan(jan):
                skipped_no_jan += 1
                continue
            params = {
                "jan": jan,
                "item_code": payload.get("item_code"),
                "upc": jan,
                "display_name": payload.get("display_name"),
                "handling_status": payload.get("handling_status"),
                "std_cost": payload.get("std_cost"),
                "avg_cost": payload.get("avg_cost"),
                "now": now,
            }
            conn.execute(sql, params)
            inserted += 1
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            errors += 1
            try:
                _record_error(conn, run_id, n, str(e), raw)
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass

    _finalize_run(conn, run_id, total=len(rows), inserted=inserted, errors=errors)
    return {
        "run_id": run_id, "total": len(rows),
        "inserted": inserted, "errors": errors,
        "period_start": None, "period_end": None,
    }


# ============================================================
# Ingestor 10：item_monthly_turnover（アイテム月完売率300）
# ============================================================
def ingest_monthly_turnover(
    path, conn, *, source_name: str | None = None
) -> dict:
    """アイテム月完売率300.xls → item_monthly_turnover (按 item_code × location × month UPSERT).

    19 列, header_row=7, period 在 row 4.
    sell_through_rate = qty_sold / (open_qty + qty_total_in)
    risk_label: ≥0.9 断货风险 / 0.5-0.9 正常 / <0.5 压库存
    """
    path = Path(path)
    source_name = source_name or path.name
    period_start, period_end = _extract_period(path)
    # year_month: 取 period_start 的 YYYYMM
    year_month = period_start.replace("-", "")[:6] if period_start else ""

    run_id = _start_run(conn, "monthly_turnover", source_name)

    # 同 (period_start, period_end) 重新上传 → 删旧再插
    if year_month:
        conn.execute(
            "DELETE FROM item_monthly_turnover WHERE year_month = ?",
            (year_month,),
        )

    rows = parse_to_dicts(path, header_row=7)

    # 预读 item_code → jan 映射 (从 item_v2)
    item_to_jan: dict[str, str] = {}
    try:
        cur = conn.execute("SELECT item_code, jan FROM item_v2 WHERE item_code IS NOT NULL")
        for row in cur.fetchall():
            ic = row[0] if not hasattr(row, "keys") else row["item_code"]
            j = row[1] if not hasattr(row, "keys") else row["jan"]
            if ic and j:
                item_to_jan[str(ic)] = str(j)
    except Exception:
        item_to_jan = {}

    sql = """
        INSERT INTO item_monthly_turnover (
            item_code, jan, location, department, year_month,
            open_qty, open_avg_cost, open_amount,
            qty_received, qty_other_in, qty_total_in,
            manual_input, last_received_at,
            qty_sold, qty_other_out, qty_total_out, out_amount, last_sold_at,
            close_qty, close_avg_cost, close_amount,
            sell_through_rate, risk_label,
            imported_at
        ) VALUES (
            :item_code, :jan, :location, :department, :year_month,
            :open_qty, :open_avg_cost, :open_amount,
            :qty_received, :qty_other_in, :qty_total_in,
            :manual_input, :last_received_at,
            :qty_sold, :qty_other_out, :qty_total_out, :out_amount, :last_sold_at,
            :close_qty, :close_avg_cost, :close_amount,
            :sell_through_rate, :risk_label,
            :imported_at
        )
    """
    now = _now_iso()
    inserted = errors = 0
    for n, raw in enumerate(rows, start=1):
        try:
            item_code = _to_str(raw.get("アイテム"))
            # 跳过分组行 (アイテム 为空 或 値是分类标签)
            if not item_code:
                continue
            if item_code in {"在庫アイテム", "合計", "総合計", "総計"}:
                continue

            location = _to_str(raw.get("場所"))
            department = _to_str(raw.get("部門"))

            open_qty = _to_float(raw.get("開始時の手持在庫数量"))
            qty_total_in = _to_float(raw.get("合計入庫数量"))
            qty_sold = _to_float(raw.get("売上"))

            # 派生 sell-through rate
            denom = (open_qty or 0.0) + (qty_total_in or 0.0)
            rate = (qty_sold or 0.0) / denom if denom > 0 else None

            if rate is None:
                risk_label = "无数据"
            elif rate >= 0.9:
                risk_label = "断货风险"
            elif rate < 0.5:
                risk_label = "压库存"
            else:
                risk_label = "正常"

            payload = {
                "item_code": item_code,
                "jan": item_to_jan.get(item_code),
                "location": location,
                "department": department,
                "year_month": year_month,
                "open_qty": open_qty,
                "open_avg_cost": _to_float(raw.get("開始平均原価")),
                "open_amount": _to_float(raw.get("開始時の手持在庫額")),
                "qty_received": _to_float(raw.get("受領")),
                "qty_other_in": _to_float(raw.get("その他の在庫入庫")),
                "qty_total_in": qty_total_in,
                "manual_input": _to_float(raw.get("入力値")),
                "last_received_at": _to_str(raw.get("前回の受領日")),
                "qty_sold": qty_sold,
                "qty_other_out": _to_float(raw.get("その他の在庫出庫")),
                "qty_total_out": _to_float(raw.get("合計出庫数量")),
                "out_amount": _to_float(raw.get("出庫価額")),
                "last_sold_at": _to_str(raw.get("前回の売上日")),
                "close_qty": _to_float(raw.get("終了時の手持在庫数量")),
                "close_avg_cost": _to_float(raw.get("期末平均原価")),
                "close_amount": _to_float(raw.get("終了時の手持在庫額")),
                "sell_through_rate": rate,
                "risk_label": risk_label,
                "imported_at": now,
            }
            conn.execute(sql, payload)
            inserted += 1
        except Exception as e:
            errors += 1
            try:
                _record_error(conn, run_id, n, str(e), raw)
            except Exception:
                pass

    _finalize_run(conn, run_id, total=len(rows), inserted=inserted, errors=errors)
    return {
        "run_id": run_id,
        "total": len(rows),
        "inserted": inserted,
        "errors": errors,
        "period_start": period_start,
        "period_end": period_end,
    }


# ============================================================
# 自动派发：根据文件名启发式选择 ingestor
# ============================================================
INGESTOR_REGISTRY: dict[str, callable] = {
    "inventory": ingest_inventory_snapshot,
    "inventory_multi": ingest_inventory_snapshot_multi,
    "asean_monthly": ingest_sales_asean_monthly,
    "asean_daily": ingest_sales_asean_daily,
    "export_item": ingest_sales_export_item,
    "export_store": ingest_sales_export_store,
    "turnover": ingest_inventory_turnover,
    "shopee_orders": ingest_shopee_orders_raw,
    "shopee_income": ingest_shopee_income,
    "item_summary": ingest_item_summary,
    "monthly_turnover": ingest_monthly_turnover,
}


def detect_ingestor(filename: str) -> str | None:
    """根据文件名启发式判断使用哪个 ingestor。"""
    n = filename
    # 新格式（多仓多级表头）：在庫のスナップショット-980 / -xxx → inventory_multi
    if "在庫のスナップショット" in n:
        return "inventory_multi"
    if "在庫数残数" in n or "通常在庫" in n:
        return "inventory"
    if "月完売率" in n or "完売率" in n:
        return "monthly_turnover"
    if "在庫回転率" in n or "回転率" in n:
        return "turnover"
    if "ASEAN" in n and "前日" in n:
        return "asean_daily"
    if "ASEAN" in n and "店舗別" in n:
        return "asean_monthly"
    if "輸出" in n and "アイテム別" in n:
        return "export_item"
    if "輸出" in n and "店舗別" in n:
        return "export_store"
    # Shopee 财务两份原表
    if "订单导出" in n or "订单导出" in n.lower():
        return "shopee_orders"
    if "income" in n.lower() and ("已拨款" in n or "拨款" in n):
        return "shopee_income"
    if "mtkshop" in n.lower() and "income" in n.lower():
        return "shopee_income"
    # NetSuite アイテム概要 (8 列, R1 表头, page 03 平均原価源)
    # 注: 需早于 export_item 检查, 因「アイテム」三字 + 数字尾缀 vs「輸出 アイテム別」二者不冲突
    if n.startswith("アイテム") or n.startswith("ｱｲﾃﾑ"):
        return "item_summary"
    return None
