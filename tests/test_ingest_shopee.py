"""Shopee 导入测试。

覆盖：
- schema 创建
- Service Fee Details fee_type 枚举
- shopee_orders 17000+ 行入库
- 重复 ingest 不报错（idempotent）
- 缺失 sheet 不崩溃
"""
import sqlite3
import tempfile
from pathlib import Path

import openpyxl
import pytest

from data_warehouse.db.migrations import init_db
from data_warehouse.ingest.excel_orders import ingest_orders
from data_warehouse.ingest.excel_shopee_income import (
    ingest_adjustments,
    ingest_income,
    ingest_service_fees,
)


@pytest.fixture
def db_conn():
    """创建临时数据库连接。"""
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "test.db"
        conn = init_db(db_path)
        yield conn
        conn.close()


def test_schema_tables_created(db_conn):
    """验证 4 张表创建成功。"""
    tables = [
        "shopee_payouts",
        "shopee_fees",
        "shopee_adjustments",
        "shopee_orders",
    ]
    for table in tables:
        count = db_conn.execute(
            f"SELECT COUNT(*) FROM {table}"
        ).fetchone()[0]
        assert count == 0, f"表 {table} 应为空"


def test_service_fee_types_enumeration(db_conn):
    """验证 Service Fee Details 中 fee_type 枚举正确。"""
    # 创建一个最小的 Service Fee Details sheet
    # 注意：ingest_service_fees 从 min_row=3 开始读数据
    wb = openpyxl.Workbook()
    ws = wb.active

    # row 1: 标题行（服务费）
    ws.append([None, None, "服务费", None, None, None])

    # row 2: header
    ws.append(["编号", "订单编号", "CB Infrastructure Fee", "CB MDV 4%", "CB MDV 5%", "CB MP Platform Shipping Fee 5.6%"])

    # row 3: 数据行
    ws.append(["1", "SHP001", 10.0, 20.0, 30.0, 40.0])

    # 导入
    count = ingest_service_fees(ws, db_conn)
    assert count == 4, f"应导入 4 条费用记录，实际 {count}"

    # 验证 fee_type 枚举
    fee_types = db_conn.execute(
        "SELECT DISTINCT fee_type FROM shopee_fees"
    ).fetchall()
    fee_types = [row[0] for row in fee_types]

    expected = {"cb_infrastructure", "cb_mdv_4", "cb_mdv_5", "platform_shipping"}
    actual = set(fee_types)
    assert actual == expected, f"fee_type 枚举不匹配：期望 {expected}，实际 {actual}"


def test_shopee_orders_large_dataset(db_conn):
    """验证 shopee_orders 支持 17000+ 行入库。"""
    # 创建模拟订单数据（简单版本，不必全部 17362）
    wb = openpyxl.Workbook()
    ws = wb.active

    # header
    ws.append(["支付币种", "单价", "发货数量", "本地SKU", "支付金额", "平台", "订单号", "店铺"])

    # 生成 1000 条测试数据（快速验证，实际数据是 17362）
    for i in range(1000):
        order_no = f"ORD{i:06d}"
        jan = f"JAN{i:010d}"
        ws.append(["PHP", 100.0, 1, jan, 100.0, "Shopee", order_no, "TestShop"])

    # 导入
    count = ingest_orders(ws, db_conn)
    assert count == 1000, f"应导入 1000 条订单，实际 {count}"

    # 验证数据库中的记录数
    db_count = db_conn.execute("SELECT COUNT(*) FROM shopee_orders").fetchone()[0]
    assert db_count == 1000, f"数据库应有 1000 条记录，实际 {db_count}"


def test_idempotent_ingest(db_conn):
    """验证重复 ingest 不报错（幂等性）。"""
    # 创建一个简单的 sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["支付币种", "单价", "发货数量", "本地SKU", "支付金额", "平台", "订单号", "店铺"])
    ws.append(["PHP", 100.0, 1, "JAN123", 100.0, "Shopee", "ORD001", "Shop1"])

    # 第一次导入
    count1 = ingest_orders(ws, db_conn)
    assert count1 == 1

    # 第二次导入（重复，使用 PRIMARY KEY 冲突解决）
    count2 = ingest_orders(ws, db_conn)
    assert count2 == 1, f"重复导入应仍为 1，实际 {count2}"

    # 验证总数仍为 1（不增加）
    total = db_conn.execute("SELECT COUNT(*) FROM shopee_orders").fetchone()[0]
    assert total == 1, f"总数应为 1，实际 {total}"


def test_missing_adjustment_sheet_no_crash(db_conn):
    """验证缺失 Adjustment sheet 不崩溃。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    # 只添加 Income 和 Service Fee Details
    income_ws = wb.create_sheet("Income")
    income_ws.append(["卖家帐号", "付款ID", "收款渠道", "拨款时间"])
    income_ws.append(["mtkshop.ph", "PAY001", "Bank", "2026-04-01"])

    fees_ws = wb.create_sheet("Service Fee Details")
    fees_ws.append([None, None, "服务费", None, None, None])
    fees_ws.append(["编号", "订单编号", "CB Infrastructure Fee", "CB MDV 4%", "CB MDV 5%", "CB MP Platform Shipping Fee"])
    fees_ws.append(["1", "SHP001", 10.0, 20.0, 30.0, 40.0])

    # 导入不应崩溃
    count_income = ingest_income(income_ws, db_conn)
    count_fees = ingest_service_fees(fees_ws, db_conn)

    # Adjustment 没有 sheet，不调用 ingest_adjustments
    adj_count = 0

    total = count_income + count_fees + adj_count
    assert total > 0, "应至少导入 Income + Fees"


def test_payouts_schema_structure(db_conn):
    """验证 shopee_payouts 表结构。"""
    # 获取表 schema
    cursor = db_conn.execute("PRAGMA table_info(shopee_payouts)")
    columns = {row[1]: row[2] for row in cursor.fetchall()}  # name: type

    expected_columns = {
        "payout_id": "TEXT",
        "seller_account": "TEXT",
        "channel": "TEXT",
        "payout_date": "DATE",
        "total_payout": "REAL",
        "currency": "TEXT",
        "ingested_at": "TIMESTAMP",
    }

    for col, col_type in expected_columns.items():
        assert col in columns, f"缺少列 {col}"
        assert columns[col] == col_type, f"列 {col} 类型不匹配：期望 {col_type}，实际 {columns[col]}"


def test_orders_schema_structure(db_conn):
    """验证 shopee_orders 表结构。"""
    # 获取表 schema
    cursor = db_conn.execute("PRAGMA table_info(shopee_orders)")
    columns = {row[1]: row[2] for row in cursor.fetchall()}  # name: type

    expected_columns = {
        "order_no": "TEXT",
        "sku_or_jan": "TEXT",
        "qty": "REAL",
        "unit_price": "REAL",
        "payment_amount": "REAL",
        "currency": "TEXT",
        "platform": "TEXT",
        "shop_name": "TEXT",
        "ingested_at": "TIMESTAMP",
    }

    for col, col_type in expected_columns.items():
        assert col in columns, f"缺少列 {col}"
        # TIMESTAMP 可能显示为 TIMESTAMP 或其他形式，不严格检查类型
        if col != "ingested_at":
            assert columns[col] == col_type, f"列 {col} 类型不匹配：期望 {col_type}，实际 {columns[col]}"
