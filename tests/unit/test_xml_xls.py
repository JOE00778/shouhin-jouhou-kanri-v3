"""测试 SpreadsheetML XML 解析器 + OOXML smart fallback。"""
from __future__ import annotations

from pathlib import Path

import pytest

from shared.xml_xls import (
    _sniff_format,
    detect_header_row,
    iter_rows,
    iter_rows_smart,
    parse_smart,
    parse_to_dicts,
)

DATA_DIR = Path("/Users/joe/CC/data")
INVENTORY_FILE = DATA_DIR / "FB全倉庫通常在庫数残数検索結果362.xls"
SALES_MONTHLY_FILE = DATA_DIR / "【ASEAN】店舗別売上　集計専用705.xls"
SALES_DAILY_FILE = DATA_DIR / "【ASEAN】店舗別売上（前日）-646.xls"
EXPORT_ITEM_FILE = DATA_DIR / "【輸出】アイテム別売上（概要）_JO-14.xls"
EXPORT_STORE_FILE = DATA_DIR / "【輸出】店舗別売上_JO-800.xls"
TURNOVER_FILE = DATA_DIR / "在庫回転率-959.xls"


def _skip_if_missing(path: Path):
    if not path.exists():
        pytest.skip(f"data file missing: {path}")


# ============================================================
# Saved Search 型（header 在第 0 行）
# ============================================================
class TestInventorySavedSearch:
    def test_inventory_first_row_is_header(self):
        _skip_if_missing(INVENTORY_FILE)
        rows = list(iter_rows(INVENTORY_FILE))
        assert rows[0][0] == "内部ID"
        assert "アイテム" in rows[0]
        assert "平均原価合計" in rows[0]
        assert "アイテム定義原価" in rows[0]

    def test_inventory_parse_to_dicts(self):
        _skip_if_missing(INVENTORY_FILE)
        records = parse_to_dicts(INVENTORY_FILE, header_row=0)
        assert len(records) > 7000
        first = records[0]
        assert first["内部ID"] == "51206"
        assert first["UPCコード"] == "7611160093868"
        assert first["平均原価合計"] is not None

    def test_detect_header_row_for_inventory(self):
        _skip_if_missing(INVENTORY_FILE)
        # 库存 saved search 表头在第 0 行
        assert detect_header_row(INVENTORY_FILE) == 0


# ============================================================
# Report 型（header 在第 6 行，前面是 preamble）
# ============================================================
class TestReportSales:
    def test_monthly_sales_header_row_6(self):
        _skip_if_missing(SALES_MONTHLY_FILE)
        rows = list(iter_rows(SALES_MONTHLY_FILE))
        # 行 0-5 是 preamble，行 6 是表头
        assert rows[6][0] == "FB_店舗"
        assert "販売数量" in rows[6]
        assert "粗利" in rows[6]

    def test_monthly_sales_parse(self):
        _skip_if_missing(SALES_MONTHLY_FILE)
        records = parse_to_dicts(SALES_MONTHLY_FILE, header_row=6)
        assert len(records) > 5000
        first = records[0]
        assert first["FB_店舗"] in ["Shopee BR", "Shopee SG", "Shopee TW", "Shopee PH"]
        assert "粗利" in first

    def test_daily_sales_parse(self):
        _skip_if_missing(SALES_DAILY_FILE)
        records = parse_to_dicts(SALES_DAILY_FILE, header_row=6)
        assert len(records) > 100

    def test_export_item_parse(self):
        _skip_if_missing(EXPORT_ITEM_FILE)
        records = parse_to_dicts(EXPORT_ITEM_FILE, header_row=6)
        assert len(records) > 1000
        # 出口アイテム別 应该有 商品ランク
        assert "商品ランク" in records[0]

    def test_export_store_parse(self):
        _skip_if_missing(EXPORT_STORE_FILE)
        records = parse_to_dicts(EXPORT_STORE_FILE, header_row=6)
        assert len(records) > 5000

    def test_turnover_parse(self):
        _skip_if_missing(TURNOVER_FILE)
        records = parse_to_dicts(TURNOVER_FILE, header_row=6)
        assert len(records) > 10000
        # 在庫回転率 应该有 回転率
        assert "回転率" in records[0]


# ============================================================
# detect_header_row 启发式
# ============================================================
class TestDetectHeader:
    def test_inventory_detected_as_row_0(self):
        _skip_if_missing(INVENTORY_FILE)
        assert detect_header_row(INVENTORY_FILE) == 0

    def test_monthly_sales_detected_as_row_6(self):
        _skip_if_missing(SALES_MONTHLY_FILE)
        assert detect_header_row(SALES_MONTHLY_FILE) == 6

    def test_export_item_detected_as_row_6(self):
        _skip_if_missing(EXPORT_ITEM_FILE)
        assert detect_header_row(EXPORT_ITEM_FILE) == 6


# ============================================================
# Smart fallback (SpreadsheetML XML + OOXML 自动派发)
# ============================================================
class TestSmartFallback:
    def test_sniff_xml_for_netsuite_export(self):
        _skip_if_missing(SALES_MONTHLY_FILE)
        assert _sniff_format(SALES_MONTHLY_FILE) == "xml"

    def test_smart_xml_path_matches_legacy(self):
        _skip_if_missing(SALES_MONTHLY_FILE)
        legacy = parse_to_dicts(SALES_MONTHLY_FILE, header_row=6)
        smart = parse_smart(SALES_MONTHLY_FILE, header_row=6)
        assert len(smart) == len(legacy)

    def test_smart_ooxml_xlsx_fallback(self, tmp_path):
        """构造一个真 OOXML .xlsx, 验证 smart 走 openpyxl 路径成功。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["jan", "qty", "name"])
        ws.append(["4901234567890", 5, "测试商品 A"])
        ws.append(["4901234567891", 12, "测试商品 B"])
        xlsx = tmp_path / "fake_export.xlsx"
        wb.save(str(xlsx))

        assert _sniff_format(xlsx) == "ooxml"
        rows = parse_smart(xlsx, header_row=0)
        assert len(rows) == 2
        assert rows[0]["jan"] == "4901234567890"
        assert rows[0]["qty"] == 5
        assert rows[1]["name"] == "测试商品 B"
